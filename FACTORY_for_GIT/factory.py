from tkinter import*
import tkinter.ttk as ttk
import csv
import pickle
import copy
from datetime import datetime
from time import monotonic
from tkinter.messagebox import *
import pyodbc
import decimal
import pyttsx3
import sqlite3
from sqlite3 import Error
from ftplib import FTP_TLS
import helpic_factory  as helpic              # for special data
import data_dict as dd
import os
from openpyxl import Workbook

d_ostatky = {}
d_selected = {}
d_zagotovka = {}
d_length_shtribs ={}
d_ostatky_kharkov = {}
d_saved_zapusk = {}
d_union_sel = {}
d_ostatky_copy_screen = {}  # for deleting row by DELETE in screen_rasschet
d_karta_detaley = {}


d_naryad = {}  # f_connection(), f_nezavershonnie_naryadi() for creating 'd_union_sel'
d_vipusk = {}  # f_connection(), f_nezavershonnie_naryadi() for creating 'd_union_sel'
d_zayavka = {}
d_rezerv = {}
l_for_save = []

text_refresh = 'Обновить'
flag_base_factory = False      # for preventing rewriting 'base_factory.db' after WRONG connection to SQL-1C-BASE
flag_ent_count = 0

# dictionary for specification in APP
d_specification = dd.d_specification

# dictionary for nomenklatura in APP
d_original = dd.d_original

#for d_oboroti 'IP54'
d_dubli_ip54 = dd.d_dubli_ip54

def f_close(event):
    f_exit()

def f_exit():
    root.destroy()

# Menu File/About
def f_about():
    showinfo('FACTORY', \
             """
             1.  При запуске приложения, загружается:
               - остатки склад Цех №2 из 1C
               - остатки склад М/к из 1C
               - обороты за период из 1C
                 (текДата - 1 мес : текДата - 13 мес). Расходные Накладные по ВСЕМ складам.
                 (Номенклатура IP54 учитывается как  IP31)
                 'ПОЛУФАБРИКАТЫ' - Выпуски Продукции
               - остатки предыдущих выпусков смотреть в 'ostatky_kharkov.txt'
               - карта деталей
             2. Information for MASTER("Верхний левый блок")
               2.1.  Кнопка "Рассчет" делит остаток за период на числоб указанное в ячейке рядом.
                     (целое числоб округление - арифметическое)
               2.2.  "двойной щелчек правой кнопки 'мыши'":
                  - производство (ввод/изменение к-ва в колонке "Производство" )
                  - спецификация (открыть спецификацию)
                  - запуск (заносит позицию в Selected items) 
             3. Selected items ("Верхний правый блок")
               3.1. Кнопка "Заготовка" заносит номенклатуру и ее к-во исходя из номенклатуры запуска и ее спецификации.
               3.2.  "двойной щелчек правой кнопки 'мыши'":
                  - колличество  (изменение к-ва в колонке "К-во")
                  - спкцификация (открыть спецификацию) 
             4. Production ("нижний левый блок")
               4.1. Кнопка "Сохранить в файл" сохраняет запуск в 'product_orders.txt'.
               4.2. Кнопка "Сохранить DELTA" сохраняет запуск в 'ostatky_kharkov.pkl'.
                   (можно просмотреть в 'ostatky_kharkov.txt')   
               4.3.  "двойной щелчек правой кнопки 'мыши'":
                  - остаток (изменяет остатки предыдущего выпуска)
                  - запуск  (изменяет 'расчитанный' запуск)  
             5. Shtribs ("нижний првый блок")
               выводит ширину и длину штрибсов для изготовления партии запуска.
             6. 'PUT M'   - записывает в текст для публикации на сайте, при следующем соидинении.
                'CHECK M' - проверяет какой текст записан и передается.
             
              Don't throw the slippers!!!
              Good luck!
         
              Kostiantyn Sh
              August_2019 - March_2020'
             """)

def f_oboroty():
    global data_tovari, d_ostatky, d_original

    d_oboroty = {}
    start = monotonic()
    print("Starting to create 'd_oboroty' for ALL SCLAD's............")

    # for count start anf finish time period for product #
    time_now = datetime.now()
    # print('time_now', time_now)
    l_time_now = list(map(int, time_now.strftime('%Y %m %d').split()))

    l_time_now[0] += 2000

    l_finish_product = [i for i in l_time_now]

    l_finish_product[0] = (l_finish_product[0] - 1) if l_finish_product[1] == 1 else l_finish_product[0]
    l_finish_product[1] = (l_finish_product[1] - 1) if l_finish_product[1] != 1 else 12

    l_start_product = [i for i in l_finish_product]

    l_start_product[0] = str(l_start_product[0] - 1)
    l_finish_product[0] = str(l_finish_product[0])
    l_start_product[1] = l_finish_product[1] = '{:02}'.format(l_finish_product[1])
    l_start_product[2] = l_finish_product[2] = '{:02}'.format(l_finish_product[2])


    start_product_text  = '-'.join(l_start_product)       # for label_period_product # root
    start_product_text  = '2' + start_product_text[1:]    # dispite +2000 SQL shift

    finish_product_text = '-'.join(l_finish_product)      # for label_period_product # root

    finish_product_text = '2' + finish_product_text[1:]  # dispite +2000 SQL shift

    finish_product = int(''.join(l_finish_product))
    start_product = int(''.join(l_start_product))

    control = 0
    for i in data_tovari:
        # (datetime.datetime(4019, 6, 20, 9, 32, 37), Decimal('1'), Decimal('5.000'), '00000004623', 'МКН 642М IP31', 'Склад М/К', [0, 0, 0, 217])
        # list(i[6])[-1] == 217 realizaciya
        # !!!!!!!!!!!! TEMPORARY SOLUTION !!!!!!!!!!

        date = int((str(i[0]).split()[0]).replace('-', ''))  # 40121231

        if start_product <= date <= finish_product and d_original.get(i[3], None) != None and list(i[6])[-1] == 217:   # list(i[6])[-1] == 217 - realizaciya

            if d_oboroty.get(i[3], None) == None:
                 d_oboroty[i[3]] = [[i[4], i[0], i[5], i[2]]]
            else:
                 l = d_oboroty[i[3]]
                 d_oboroty[i[3]] = l + [[i[4], i[0], i[5], i[2]]]

        # for NOMENKLATURA with 'IP54'
        elif start_product <= date <= finish_product  and d_dubli_ip54.get(i[3], None) != None and list(i[6])[-1] == 217:

            if d_oboroty.get(d_dubli_ip54.get(i[3], None)[0], None) == None:
                 d_oboroty[d_dubli_ip54.get(i[3], None)[0]] = [[i[4], i[0], i[5], i[2]]]
            else:
                 l = d_oboroty[d_dubli_ip54.get(i[3], None)[0]]
                 d_oboroty[d_dubli_ip54.get(i[3], None)[0]] = l + [[i[4], i[0], i[5], i[2]]]

        # for nomenklatura POLUFABRIKATI - oboroti==vipuski
        # list(i[6])[-1] == 232 and int(i[1]) == 0 - PRIHOD /vipuski/
        # list(i[6])[-1] == 232 and int(i[1]) == 1 - RASHOD /peremescheniya po naryadu, spisanie /

        # for adding RASHOD /peremescheniya po naryadu, spisanie /  2020.16.06
        elif start_product <= date <= finish_product  and d_original.get(i[3], None) != None and list(i[6])[-1] == 232 and int(i[1]) == 1:
            if d_oboroty.get(i[3], None) == None:
                 d_oboroty[i[3]] = [[i[4], i[0], i[5], i[2]]]

            else:
                 l = d_oboroty[i[3]]
                 d_oboroty[i[3]] = l + [[i[4], i[0], i[5], i[2]]]

        # checker for control
        # if 40190101 <= date <= 40191231 and d_original.get(i[3], None) != None and list(i[6])[-1] == 217:                     # list(i[6])[-1] == 217 - realizaciya
        # if 40190101 <= date <= 40191231 and d_original.get(i[3], None) != None and list(i[6])[-1] == 232 and int(i[1]) == 1:  # peremescheniya po naryadu, spisanie
        #     if i[4] == 'МПС 126 Панель монтажная  (оц)':
        #         control += i[2]
        #         print(control, i)



    # 'd_oboroty'
    # 00000014854 : [['МКС 1464 IP54 (без МП)', datetime.datetime(4018, 9, 4, 11, 46, 17), 'Склад М/К', Decimal('1.000')],
    # ['МКС 1464 IP31 (без МП)', datetime.datetime(4018, 9, 4, 11, 44, 11), 'Склад М/К', Decimal('1.000')],
    # ['МКС 1464 IP31 (без МП)', datetime.datetime(4018, 11, 27, 15, 46, 23), 'Склад Цех №2', Decimal('1.000')],
    # ['МКС 1464 IP31 (без МП)', datetime.datetime(4018, 9, 14, 10, 4, 58), 'Склад М/К', Decimal('1.000')],
    # ['МКС 1464 IP31 (без МП)', datetime.datetime(4018, 9, 27, 0, 0), 'Склад М/К', Decimal('1.000')],
    # ['МКС 1464 IP31 (без МП)', datetime.datetime(4018, 6, 25, 0, 0), 'Склад М/К', Decimal('1.000')],
    # ['МКС 1464 IP31 (без МП)', datetime.datetime(4018, 10, 2, 11, 6, 18), 'Склад М/К', Decimal('1.000')],
    # ['МКС 1464 IP31 (без МП)', datetime.datetime(4018, 12, 20, 9, 6, 50), 'Склад М/К', Decimal('1.000')],
    # ['МКС 1464 IP31 (без МП)', datetime.datetime(4018, 7, 11, 16, 19, 1), 'Склад М/К', Decimal('1.000')],
    # ['МКС 1464 IP31 (без МП)', datetime.datetime(4018, 10, 19, 0, 0), 'Склад М/К', Decimal('2.000')],
    # ['МКС 1464 IP54 (без МП)', datetime.datetime(4018, 7, 27, 8, 42, 55), 'Склад М/К', Decimal('1.000')]]

    # checking 'd_oboroty'
    # for i in d_oboroty:
    #     if d_oboroty.get(i, None)[0][0] == 'ДВ 126М.Дверь.Д':
    #         print(i, d_oboroty.get(i, None))

    d_sel = {}
    for i in d_oboroty:
        d_sel[i] = [d_original.get(i)[0], sum([i[3] for i in d_oboroty.get(i, None)])]  # 'd_oboroty[i][0][0]' - nomenklatura

    # for checking 'd_sel'
    # for i in d_sel:
    #     print(i, d_sel.get(i, None))

    for i in d_sel:
        #  00000011696 ['ЕР 1884/2 В', Decimal('8.000')]
       for keys in d_ostatky:
           if i == keys:
               l = d_ostatky[keys]
               l.append(int(d_sel[i][1]))
               d_ostatky[keys] = l

    print("\033[0m 'd_oboroty' for ALL SCLAD's is created \033[32m{: .5f} seconds \033[0m".format(monotonic()-start))
    label_period_product.configure(text='ПРОДАЖИ: {}  {}'.format(start_product_text, finish_product_text))

# # adding '' for columns product, store, zapusk
def f_add_pr_st_zapusk():
    # d_ostatky_kharkov {nomenklatura: quantity}
    # d_ostatky_kharkov = {'БС 206М.Боковая стенка.Д': 1, 'Дв 208М.ЕР.Дверь.Д': 1, 'ЗС 208М.ЕР.Задняя стенка.Д': 1}
    # d_ostatky {Code: [group, nomenklatura, quantity_mk, quantity_ceh}
    for keys, values in d_ostatky.items():
        l = d_ostatky[keys]
        l_new = l[:4]

        # d_ostatky {Code: [group, nomenklatura, quantity_mk, quantity_ceh, PRODUCT}

        try:
            naryad = 0
            vipusk = 0
            for data in  d_union_sel.get(l[1], None):
                naryad += data[0][3]
                if data[1] == '':
                    vipusk += 0
                else:
                    for ii in data[1]:
                        vipusk += ii[3]

                ostatok_nar = naryad - vipusk

        except TypeError:
            ostatok_nar = ''

        l_new.append(ostatok_nar)

        # for nomenklatura without 'oboroty'
        try:
            l_new.append(l[4])
        except IndexError:
            l_new.append('')
        # d_ostatky {Code: [group, nomenklatura, quantity_mk, quantity_ceh, PRODUCT, OBOROTY=''}

        # for adding ZAPUSK
        l_new.append('')
        # d_ostatky {Code: [group, nomenklatura, quantity_mk, quantity_ceh, PRODUCT, OBOROTY='', ZAPUSK = ''}
        d_ostatky[keys] = l_new

        # for adding SHTRIBS
        # for NOMENKLATURA with SHTRIBS in karta_detaley
        # d_karta_detaley {nomenklatura: [place, long, width, shtribs, x, y]}
        nomenklatura = d_ostatky[keys][1]
        if d_karta_detaley.get(nomenklatura, None) != None:
            d_ostatky[keys]  += [d_karta_detaley.get(nomenklatura, None)[3]]
        else:
            d_ostatky[keys]  += ['']
        # d_ostatky {Code: [group, nomenklatura, quantity_mk, quantity_ceh, PRODUCT, OBOROTY='', ZAPUSK = '', SHTRIBS=''}

        # for adding OSTATKY CEH
        if d_ostatky_kharkov.get(nomenklatura, None) != None:
            d_ostatky[keys] += [d_ostatky_kharkov.get(nomenklatura, None)]
        else:
            d_ostatky[keys] += ['']
        # d_ostatky {Code: [group, nomenklatura, quantity_mk, quantity_ceh, PRODUCT, OBOROTY='', ZAPUSK = '', SHTRIBS='', 'OSTATKY CEH'}

    # FOR  CONTROLL
    # print(d_union_sel) #
    # for i in d_union_sel:
    #     try:
    #         print(d_union_sel.values()[1:])
    #     except:
    #         None

    # for checking d_ostatky #
    # for i in d_ostatky:
    #     print(i, d_ostatky[i])

# creating table 'tree_info' - MAIN TABLE
def f_table_tree_inf():
    global d_ostatky, d_zayavka, d_rezerv

    # 'd_zayavka'
    # 00000012162 [[datetime.datetime(4019, 9, 4, 9, 25, 49), 'ЕТР-КОМПЛЕКС', Decimal('2.000'), '03.09.2019', 'М 4.75 В '],
    # [datetime.datetime(4019, 9, 3, 15, 13, 2), 'КМ-РІШЕННЯ', Decimal('8.000'), '03.09.2019', 'М 4.75 В '],
    # [datetime.datetime(4019, 9, 5, 15, 31, 43), 'Мадек', Decimal('20.000'), '', 'М 4.75 В '],
    # [datetime.datetime(4019, 8, 29, 13, 57, 31), 'СТИКС-ОИЛ ИНЖИНИРИНГ', Decimal('4.000'), '29.08.2019', 'М 4.75 В ']]

    # 'd_ostatky'
    #  00000013144 ['Полуфабрикаты', 'Пластина боковая ПлБ41Т.Д', '', -120, '', '', '', '']
    # d_ostatky {Code: [group, nomenklatura, quantity_mk, quantity_ceh, PRODUCT, OBOROTY='', ZAPUSK = '', SHTRIBS='', 'OSTATKY CEH'}
    # !!! 'OSTATKY CEH' !!!  from ostatky_kharkov.txt
    # 'd_rezerv'
    # 00000015596[[datetime.datetime(4019, 9, 4, 9, 1, 21), 'Промавтоматика-Вінниця', Decimal('4.00'), 'СМГ 18.35 (1,5)'],
    #             [datetime.datetime(4019, 9, 4, 15, 39, 11), 'СОЮЗ-СВІТЛО УКРАЇНА', Decimal('6.00'), 'СМГ 18.35 (1,5)'],
    #             [datetime.datetime(4019, 9, 2, 8, 20), 'ТЕХНОТОН ЕНЕРГО', Decimal('38.00'), 'СМГ 18.35 (1,5)'],
    #             [datetime.datetime(4019, 9, 5, 16, 25, 36), 'Українське електрообладнання та інсталяція', Decimal('2.00'), 'СМГ 18.35 (1,5)']]

    # deleting rows before laoding data
    rows = tree_inf.get_children()
    for item in rows:
        tree_inf.delete(item)

    # for lines where quantity in 'zayavka' is present
    tree_inf.tag_configure('zayavka', background='lightblue', foreground='red')

    # group 'ВСЯ ПРОДУКЦИЯ'
    folder_all = tree_inf.insert('', 0, text='ВСЯ ПРОДУКЦИЯ', values=())
    folder_shtribs = tree_inf.insert('', 1, text='ШТРИБС', values=())
    counter = 0

    # l_ostatky_sorted = ['00000011755', '00000010759', '00000011731', '00000011730', '00000011728',
    l_ostatky_sorted = sorted(d_ostatky, key=lambda x: d_ostatky.get(x)[1])
    l_ostatky_sorted = list(filter(lambda x: x not in d_dubli_ip54, l_ostatky_sorted))

    # Exception 'Полуфабрикаты' from 'ВСЯ ПРОДУКЦИЯ'
    for i in sorted(d_ostatky, key=lambda x: d_ostatky.get(x)[1]):
        if d_ostatky.get(i)[0] != 'Полуфабрикаты': # exclude  'Полуфабрикаты'

            nomenklatura = d_ostatky[i][1]
            quantity_mk  = d_ostatky[i][2]
            quantity_ceh = d_ostatky[i][3]
            product      = d_ostatky[i][4]
            ostatok_ceh  = d_ostatky[i][8]

            ostatok = (0 if d_ostatky[i][2]  == '' else int(d_ostatky[i][2])) +\
                      ((0 if d_ostatky[i][3] == '' else int(d_ostatky[i][3])) if ostatok_ceh == '' else ostatok_ceh) +\
                      (0 if d_ostatky[i][4]  == '' else int(d_ostatky[i][4]))

            ostatok = '' if ostatok == 0 else ostatok

            try:
                oboroty = d_ostatky[i][5]
            except IndexError:
                oboroty = ''

            zayavka = sum([int(iii[2]) for iii in d_zayavka.get(i, None)]) if d_zayavka.get(i, None) != None else ''
            rezerv = sum([int(iii[2]) for iii in d_rezerv.get(i, None)]) if d_rezerv.get(i, None) != None else ''

            zapusk = (0 if oboroty == '' else int(oboroty)) - (0 if ostatok == '' else int(ostatok))
            zapusk = '' if zapusk  <=  0 else zapusk

            if zayavka == '':
                tree_inf.insert(folder_all, counter, text=nomenklatura, values=(quantity_mk, quantity_ceh, ostatok_ceh, product, ostatok, oboroty, zayavka, rezerv,  zapusk))
            else:
                tree_inf.insert(folder_all, counter, text=nomenklatura,\
                                values=(quantity_mk, quantity_ceh, ostatok_ceh, product, ostatok, oboroty, zayavka, rezerv, zapusk), tags='zayavka')
            counter +=1

    # groups with PLUS
    l_group = list(set(i[0] for i in d_ostatky.values())) # the list of 'groups'

    counter_folder = 1
    for i in sorted(l_group):
        folder = tree_inf.insert('', counter_folder, text=str(i), values=())
        counter_folder += 1
        for ii in l_ostatky_sorted:
            counter_row = 0
            if d_ostatky[ii][0] == i:

                nomenklatura =  d_ostatky[ii][1]
                quantity_mk =  d_ostatky[ii][2]
                quantity_ceh =  d_ostatky[ii][3]
                product =  d_ostatky[ii][4]
                ostatok_ceh = d_ostatky[ii][8]
                ostatok = (0 if  d_ostatky[ii][2] == '' else int( d_ostatky[ii][2])) + \
                      ((0 if d_ostatky[ii][3] == '' else int(d_ostatky[ii][3])) if ostatok_ceh == '' else ostatok_ceh) +\
                      (0 if  d_ostatky[ii][4] == '' else int( d_ostatky[ii][4]))
                ostatok = '' if ostatok == 0 else ostatok

                try:
                    oboroty = d_ostatky[ii][5]
                except IndexError:
                    oboroty = ''
                zayavka = sum([int(iii[2]) for iii in d_zayavka.get(ii, None)]) if d_zayavka.get(ii, None) != None else ''
                rezerv = sum([int(iii[2]) for iii in d_rezerv.get(ii, None)]) if d_rezerv.get(ii, None) != None else ''

                zapusk = (0 if oboroty == '' else int(oboroty)) - (0 if ostatok == '' else int(ostatok))
                zapusk = '' if zapusk <= 0 else zapusk

                if zayavka == '':
                    tree_inf.insert(folder, counter, text=nomenklatura,
                                    values=(quantity_mk, quantity_ceh, ostatok_ceh, product, ostatok, oboroty, zayavka, rezerv, zapusk))
                else:
                    tree_inf.insert(folder, counter, text=nomenklatura, \
                                    values=(quantity_mk, quantity_ceh, ostatok_ceh, product, ostatok, oboroty, zayavka, rezerv, zapusk),
                                    tags='zayavka')
                counter += 1

    # groups with SHTRIBS
    # d_ostatky {Code: [group, nomenklatura, quantity_mk, quantity_ceh, PRODUCT, OBOROTY='', ZAPUSK = '', SHTRIBS='', OSTATOK_CEH = ''}

    l_group_shtribs = list(set(i[7] for i in d_ostatky.values() if i[7] != ''))  # the list of 'groups shtribs'

    counter_folder_shtribs = 1

    for i in sorted(l_group_shtribs):
        folder = tree_inf.insert(folder_shtribs, counter_folder_shtribs, text=str(i), values=())
        counter_folder_shtribs += 1
        counter_row = 0
        for ii in l_ostatky_sorted:
            if d_ostatky[ii][7] == i:
                nomenklatura = d_ostatky[ii][1]
                quantity_mk = d_ostatky[ii][2]
                quantity_ceh = d_ostatky[ii][3]
                product = d_ostatky[ii][4]
                ostatok_ceh = d_ostatky[ii][8]
                ostatok = (0 if d_ostatky[ii][2] == '' else int(d_ostatky[ii][2])) + \
                          ((0 if d_ostatky[ii][3] == '' else int(d_ostatky[ii][3])) if ostatok_ceh == '' else ostatok_ceh) +\
                          (0 if d_ostatky[ii][4] == '' else int(d_ostatky[ii][4]))
                ostatok = '' if ostatok == 0 else ostatok

                try:
                    oboroty = d_ostatky[ii][5]
                except IndexError:
                    oboroty = ''
                zayavka = sum([int(iii[2]) for iii in d_zayavka.get(ii, None)]) if d_zayavka.get(ii, None) != None else ''
                rezerv = sum([int(iii[2]) for iii in d_rezerv.get(ii, None)]) if d_rezerv.get(ii, None) != None else ''

                zapusk = (0 if oboroty == '' else int(oboroty)) - (0 if ostatok == '' else int(ostatok))
                zapusk = '' if zapusk <= 0 else zapusk

                if zayavka == '':
                    tree_inf.insert(folder, counter_row, text=nomenklatura,
                                    values=(quantity_mk, quantity_ceh, ostatok_ceh, product, ostatok, oboroty, zayavka, rezerv, zapusk))
                else:
                    tree_inf.insert(folder, counter_row, text=nomenklatura, \
                                    values=(quantity_mk, quantity_ceh, ostatok_ceh, product, ostatok, oboroty, zayavka, rezerv, zapusk),
                                    tags='zayavka')
                counter_row += 1

# rewrihting "ЗАПАС/"
def f_store_ent(event):                       # for 'ent_count'
    f_store()
    return 0

def f_store():             # for 'btn_count'
    global d_ostatky, d_zayavka, d_rezerv

    d_ostatky_div = copy.deepcopy(d_ostatky)          # DEEP COPY !!!!!!!
    l_ostatky_sorted = sorted(d_ostatky_div, key=lambda x: d_ostatky.get(x)[1], reverse=True)

    try:
        divisior = int(ent_count.get()) if  int(ent_count.get()) > 0 else 1
    except ValueError:
        divisior = 1

    for i in d_ostatky_div:
        l = d_ostatky_div.get(i)
        try:
           l.append(round(int(l.pop())/int(divisior)))
        except ValueError:
            l.append('')
        d_ostatky_div[i] = l

        # !!! Optimize LATER!!!
        # deleting rows before laoding data
        rows = tree_inf.get_children()
        for item in rows:
            tree_inf.delete(item)

        # group 'ВСЯ ПРОДУКЦИЯ'
        folder_all = tree_inf.insert('', 0, text='ВСЯ ПРОДУКЦИЯ', values=())
        folder_shtribs = tree_inf.insert('', 1, text='ШТРИБС', values=())

        counter = 0
        for i in l_ostatky_sorted:
            # Exception 'Полуфабрикаты' from 'ВСЯ ПРОДУКЦИЯ'
            if d_ostatky.get(i)[0] != 'Полуфабрикаты':
                nomenklatura = d_ostatky_div[i][1]
                quantity_mk = d_ostatky_div[i][2]
                quantity_ceh = d_ostatky_div[i][3]
                product = d_ostatky[i][4]
                ostatok_ceh = d_ostatky[i][8]
                ostatok = (0 if d_ostatky_div[i][2] == '' else int(d_ostatky_div[i][2])) + \
                      ((0 if d_ostatky[i][3] == '' else int(d_ostatky[i][3])) if ostatok_ceh == '' else ostatok_ceh) +\
                      (0 if d_ostatky_div[i][4] == '' else int(d_ostatky_div[i][4]))
                ostatok = '' if ostatok == 0 else ostatok
                oboroty = round(d_ostatky_div[i][5]/divisior) if type(d_ostatky_div[i][5])==int else ''
                zayavka = sum([int(iii[2]) for iii in d_zayavka.get(i, None)]) if d_zayavka.get(i, None) != None else ''
                rezerv = sum([int(iii[2]) for iii in d_rezerv.get(i, None)]) if d_rezerv.get(i, None) != None else ''

                zapusk = (0 if oboroty == '' else int(oboroty)) - (0 if ostatok == '' else int(ostatok))
                zapusk = '' if zapusk <= 0 else zapusk

                if zayavka == '':
                    tree_inf.insert(folder_all, counter, text=nomenklatura,
                                    values=(quantity_mk, quantity_ceh, ostatok_ceh, product, ostatok, oboroty, zayavka, rezerv, zapusk))
                else:
                    tree_inf.insert(folder_all, counter, text=nomenklatura, \
                                    values=(quantity_mk, quantity_ceh, ostatok_ceh, product, ostatok, oboroty, zayavka, rezerv, zapusk),
                                    tags='zayavka')
                counter += 1

        # groups with PLUS
        l_group = list(set(i[0] for i in d_ostatky_div.values()))  # the list of 'groups'

        counter_folder = 1
        for i in sorted(l_group):
            folder = tree_inf.insert('', counter_folder, text=str(i), values=())
            counter_folder += 1
            counter_row =0

            for keys, values in d_ostatky_div.items():
                if values[0] == i:
                    nomenklatura = values[1]
                    quantity_mk = values[2]
                    quantity_ceh = values[3]
                    product = values[4]
                    ostatok_ceh = values[8]
                    ostatok = (0 if values[2] == '' else int(values[2])) + \
                          ((0 if values[3] == '' else int(values[3])) if ostatok_ceh == '' else ostatok_ceh) +\
                          (0 if values[4] == '' else int(values[4]))
                    ostatok = '' if ostatok == 0 else ostatok
                    oboroty = round(values[5]/divisior) if type(values[5])==int else ''
                    zayavka = sum([int(iii[2]) for iii in d_zayavka.get(keys, None)]) if d_zayavka.get(keys, None) != None else ''

                    rezerv = sum([int(iii[2]) for iii in d_rezerv.get(i, None)]) if d_rezerv.get(i, None) != None else ''

                    zapusk = (0 if oboroty == '' else int(oboroty)) - (0 if ostatok == '' else int(ostatok))
                    zapusk = '' if zapusk <= 0 else zapusk

                    if zayavka == '':
                        tree_inf.insert(folder, counter_row, text=nomenklatura,
                                        values=(quantity_mk, quantity_ceh, ostatok_ceh, product, ostatok, oboroty, zayavka, rezerv, zapusk))
                    else:
                        tree_inf.insert(folder, counter_row, text=nomenklatura, \
                                        values=(quantity_mk, quantity_ceh, ostatok_ceh, product, ostatok, oboroty, zayavka, rezerv, zapusk),
                                        tags='zayavka')
                    counter += 1
            # groups with SHTRIBS
            # d_ostatky {Code: [group, nomenklatura, quantity_mk, quantity_ceh, PRODUCT, OBOROTY='', ZAPUSK = '', SHTRIBS=''}

        l_group_shtribs = list(set(i[7] for i in d_ostatky.values() if i[7] != ''))  # the list of 'groups shtribs'

        counter_folder_shtribs = 1
        for i in sorted(l_group_shtribs):
            folder = tree_inf.insert(folder_shtribs, counter_folder_shtribs, text=str(i), values=())
            counter_folder_shtribs += 1
            for ii in l_ostatky_sorted:
                counter_row = 0
                if d_ostatky[ii][7] == i:

                    nomenklatura = d_ostatky[ii][1]
                    quantity_mk = d_ostatky[ii][2]
                    quantity_ceh = d_ostatky[ii][3]
                    product = d_ostatky[ii][4]
                    ostatok_ceh = d_ostatky[ii][8]
                    ostatok = (0 if d_ostatky[ii][2] == '' else int(d_ostatky[ii][2])) + \
                              ((0 if d_ostatky[ii][3] == '' else int(d_ostatky[ii][3])) if ostatok_ceh == '' else ostatok_ceh) + \
                              (0 if d_ostatky[ii][4] == '' else int(d_ostatky[ii][4]))
                    ostatok = '' if ostatok == 0 else ostatok

                    oboroty = round(d_ostatky[ii][5] / divisior) if type(d_ostatky[ii][5]) == int else ''

                    zayavka = sum([int(iii[2]) for iii in d_zayavka.get(ii, None)]) if d_zayavka.get(ii, None) != None else ''
                    rezerv = sum([int(iii[2]) for iii in d_rezerv.get(ii, None)]) if d_rezerv.get(ii, None) != None else ''

                    zapusk = (0 if oboroty == '' else int(oboroty)) - (0 if ostatok == '' else int(ostatok))
                    zapusk = '' if zapusk <= 0 else zapusk

                    if zayavka == '':
                        tree_inf.insert(folder, counter_row, text=nomenklatura, values=\
                        (quantity_mk, quantity_ceh, ostatok_ceh, product, ostatok, oboroty, zayavka, rezerv, zapusk))
                    else:
                        tree_inf.insert(folder, counter_row, text=nomenklatura,values=\
                            (quantity_mk, quantity_ceh, ostatok_ceh, product, ostatok, oboroty, zayavka, rezerv, zapusk),
                                        tags='zayavka')
                    counter_row += 1

        btn_count.configure(text='ПРОДАЖИ /{}'.format(divisior))

def f_shift_year(data):
    data = data.strftime("%Y-%m-%d %H:%M:%S")
    year = int(data[:4]) - 2000
    return str(year) + data[4:]

# common window for product, zayavka, rezerv
def f_win(win_title, tuple_columns, tree_place):
    # full tuple columns ['Дата', 'Контрагент', 'К-во', 'Производство' , 'Примечание']
    item_win = tree_place.item(tree_place.selection())
    item_text = item_win['text']

    # window PRODUCT
    window = Toplevel(root, bg='lightblue', bd=5, relief=SUNKEN)
    window.title(win_title)
    window.geometry('650x500')

    lab_window_title = MyLabel(window, text=item_text)

    tree_window = ttk.Treeview(window)
    tree_window['height'] = 15

    tree_window['column'] = tuple_columns

    # common columns
    tree_window.column('Дата', width=150,anchor='n')
    tree_window.column('К-во', width=80, anchor='n')
    tree_window.heading('Дата', text='Дата')
    tree_window.heading('К-во', text='К-во')

    # differwent columns
    if 'Производство' in tuple_columns:
        tree_window.column('Производство', width=150, anchor='n')
        tree_window.heading('Производство', text='Производство')
    if 'Контрагент' in tuple_columns:
        tree_window.column('Контрагент', width=100, anchor='e')
        tree_window.heading('Контрагент', text='Контрагент')
    if 'Примечание' in tuple_columns:
        tree_window.column('Примечание', width=100, anchor='n')
        tree_window.heading('Примечание', text='Примечание')

    lab_window_title.grid(row=1, column=0)
    tree_window.grid(row=2, column=0)
    # 'Производство'
    if win_title == 'Производство':
        # data = [[[datetime.datetime(4020, 1, 14, 10, 29, 31), '00000000235', None, Decimal('8'), 0], []]]
        data = d_union_sel.get(item_text, None)

       # preventing clicking on the FOLDER (not on the item)
        try:
           counter_folder = 1
           for i in sorted(data, key = lambda x: x[0]):
                data_win = f_shift_year(i[0][0])          # shift -2000 years
                ost_from_nar = i[0][3] - i[0][4]
                folder = tree_window.insert('', counter_folder, text=i[0][1], values=(data_win, i[0][3], ost_from_nar ))
                counter_folder += 1
                counter = 1
                for vipuski in  i[1]:
                    data_win = f_shift_year(vipuski[0])    # shift -2000 years
                    tree_window.insert(folder, counter, text=vipuski[1], values=(data_win, '', vipuski[3]))
                    counter += 1
        except TypeError:
            window.destroy()  # close product_window
            return None
    # 'Заявка'
    elif win_title == 'Заявка':
        l_zayavka_sel = []
        for i in d_zayavka:
            if d_zayavka[i][0][4] == item_text:  # FIRST LIST in the COMPAUNDED LIST
                l_zayavka_sel = d_zayavka.get(i, None)
                break

        # preventing clicking on the FOLDER (not on the item) or item with EMPTY CELL
        if len(l_zayavka_sel) > 0:
            counter = 1
            for i in sorted(l_zayavka_sel, key=lambda x: x[0]):
                data_win = f_shift_year(i[0])  # shift -2000 years
                kontragent = i[1]
                quantity = int(i[2])
                primechanie = i[3]

                tree_window.insert('', counter, text='', values=(data_win, kontragent, quantity, primechanie))
                counter += 1
        else:
            window.destroy()  # close zayavka_window
    # 'Резерв'
    elif win_title == 'Резерв':
        l_rezerv_sel = []
        for i in d_rezerv:
            if d_rezerv[i][0][3] == item_text:  # FIRST LIST in the COMPAUNDED LIST
                l_rezerv_sel = d_rezerv.get(i, None)
                break

        # preventing clicking on the FOLDER (not on the item) or item with EMPTY CELL
        if len(l_rezerv_sel) > 0:
            counter = 1
            for i in sorted(l_rezerv_sel, key=lambda x: x[0]):
                data_win = f_shift_year(i[0])  # shift -2000 years
                kontragent = i[1]
                quantity = int(i[2])

                tree_window.insert('', counter, text='', values=(data_win, kontragent, quantity))
                counter += 1

        else:
            window.destroy()  # close rezerv_window
    window.focus()
    window.bind('<Escape>',lambda x: window.destroy())

# function depends on radiobutton
def f_double_click_inf(event):
    k_var = var.get()  # for several variants on the 'table'
    if var.get() == 1:               # product
        f_win('Производство', ('Дата', 'К-во', 'Производство'), tree_inf)
    elif var.get() == 9:             # zayavka
        f_win('Заявка', ('Дата','Контрагент', 'К-во'), tree_inf)
    elif var.get() == 10:             # rezerv
        f_win('Резерв', ('Дата', 'Контрагент', 'К-во', 'Примечание'), tree_inf)
    elif var.get() == 2:             # specification
        f_win_specif(tree_inf)
    elif var.get() == 3:             # zapusk
        f_win_zapusk(tree_inf, k_var)
    elif var.get() == 11:            # ostatky ceh     !!! Is not used !!!
        pass

def f_double_click_middle(event):
    k_var = var.get()  # for several variants on the 'table'
    if var.get() == 5:      #specification
         # f_win_specif_middle()
         f_win_specif(tree_middle)
    elif var.get() == 4:
        # f_win_quantity()
        f_win_zapusk(tree_middle, k_var)


def f_double_click_bottom(event):
    k_var = var.get()  # for several variants on the 'table'
    if var.get() == 6:                 # ostatok ceh
        # f_win_quantity_bottom()
        f_win_zapusk(tree_bottom, k_var)
    elif var.get() == 7:               # zapusk bottom
        # f_win_zapusk_bottom()
        f_win_zapusk(tree_bottom, k_var)
    elif var.get() == 8:               # product bottom
        # f_win_product_bottom()
        f_win('Производство', ('Дата', 'К-во', 'Производство'), tree_bottom)

def f_double_click_screen(event):
    k_var = var_screen_z.get()  # for several variants on the 'table'
    if var_screen_z.get() == 11:   # ostatok screen
         # f_win_quantity_screen()
         f_win_zapusk(tree_screen, k_var)
    elif var_screen_z.get() == 12: # zapusk screen
        # f_win_zapusk_screen()
        f_win_zapusk(tree_screen, k_var)
    elif var_screen_z.get() == 16: # product screen
        # f_win_product_screen()
        f_win('Производство', ('Дата', 'К-во', 'Производство'), tree_screen)

def f_double_click_rasschet(event, *args):
    k_var = var.get()                # for several variants on the 'table'
    if var.get() == 13:              # ostatok ceh
       f_win_zapusk(tree_screen_rasschet, k_var, *args)

def f_win_zapusk(tree_place_zapusk, k_var, *args):            # WINDOW for changing quontity
    item_product = tree_place_zapusk.item(tree_place_zapusk.focus())
    item_text =  item_product['text']

    # preventing clicking on the FOLDER (not on the item)
    try:
        # item_product['values'] = [15, 17, '', 22, -5, 5, 587, '1\xa0388', 587, 1, 1, 6940, 0, 5]
        if k_var == 7 or k_var == 12:  # zapusk /bottom+screen/
            # item_product['values'] = [15, 17, '', 22, -5, 5, 587, '1\xa0388', 587, 1, 1, 6940, 0, 5]
            item_quantity  =  int(item_product['values'][5])

        elif k_var == 3:                                            # zapusk/inf
            item_quantity = item_product['values'][-1]

        elif k_var == 13:                                           # ostatok CEH/tree_screen_rasschet
            item_quantity = 0 if not item_product['values'][6] else int(item_product['values'][6])

        else:
            item_quantity = item_product['values'][1]               # ostatok CEH/bottom+screen
    except IndexError:
        return None

    # window PRODUCT
    zapusk = Toplevel(root, bg='lightblue',bd=5, relief=SUNKEN)
    # window title
    if k_var in [3, 7, 12]:
         zapusk.title('ZAPUSK')
         text = 'В запуск'
    elif k_var in [6, 11, 13]:
        zapusk.title('OSTATOK CEH')
        text = 'Остаток Цех №2'
    else:
        zapusk.title('QUANTITY')
        text = 'Колличество'

    zapusk.geometry('450x150')
    lab_zapusk_title = MyLabel(zapusk, text=text)
    lab_zapusk_item = MyLabel(zapusk, text=item_text)
    lab_zapusk_quantity = MyLabel(zapusk, text=item_quantity)
    ent_zapusk = Entry(zapusk, width=4, bd=5, font='arial 12', fg='black', relief=SUNKEN)
    ent_zapusk.focus()
    but_zapusk = MyButton(zapusk,text='SAVE')

    lab_zapusk_title.grid(row=1, column=0, columnspan=3, padx=5, pady=5)
    lab_zapusk_item.grid(row=2, column=0, columnspan=2, padx=5, pady=5)
    ent_zapusk.grid(row=2, column=3, padx=5, pady=5)
    lab_zapusk_quantity.grid(row=2, column=4, padx=5, pady=5)
    but_zapusk.grid(row=3, column=0, columnspan=4, padx=5, pady=5)

    but_zapusk.bind('<Button-1>', lambda event, f1=item_product, f2=ent_zapusk, f3=zapusk, f4=tree_place_zapusk, f5=k_var, f6=args: f_save_zapusk(event, f1, f2, f3, f4, f5, f6))
    ent_zapusk.bind('<Return>',   lambda event, f1=item_product, f2=ent_zapusk, f3=zapusk, f4=tree_place_zapusk, f5=k_var, f6=args: f_save_zapusk(event, f1, f2, f3, f4, f5, f6))


def f_save_zapusk(event, item_product, ent_zapusk, zapusk, tree_place_zapusk, k_var, *args):

        global d_selected                    # from 'tree_inf' for creating 'selected items' in 'table_middle'
        # d_zagotovka {nomenklatura: [need, zapusk]}
        global d_zagotovka                   # from 'tree_bottom' for changing 'd_zagotovka'
        global d_ostatky_copy_screen         # from 'tree_screen_rasschet' for changing 'd_ostatky_copy_screen'
        global l_position                    # for 'lightening' selected row in 'screen rasschet'

        # 'd_ostatky'
        #  00000013144 ['Полуфабрикаты', 'Пластина боковая ПлБ41Т.Д', '', -120, '', '', '', '']
        # d_ostatky {Code: [group, nomenklatura, quantity_mk, quantity_ceh, PRODUCT, OBOROTY='', ZAPUSK = '', SHTRIBS='', 'OSTATKY CEH'}
        global fra_rasschet

        item_text = item_product['text']

        # control inputing number
        try:
            # item_quantity = int(ent_zapusk.get())
                item_quantity = int(ent_zapusk.get())
            # if item_quantity >= 0:                       # '>= 0'
                if tree_place_zapusk == tree_inf or tree_place_zapusk == tree_middle:                             # zapusk /inf+middle/
                    d_selected[item_text] = item_quantity
                    f_table_middle()                     # creating table 'selected items'
                elif k_var in [7, 12] and (tree_place_zapusk == tree_bottom or tree_place_zapusk == tree_screen): # zapusk /bottom+screen/
                        l = d_zagotovka.get(item_text)
                        l[1] = item_quantity
                        d_zagotovka[item_text] = l
                        f_zagotovka()                    # creating table 'selected items'

                        if k_var == 12:
                            f_view_screen()              # creating table SCREEN
                elif k_var in [6, 11] and (tree_place_zapusk == tree_bottom or tree_place_zapusk == tree_screen):    # ostatok CEH
                    l = d_zagotovka.get(item_text)
                    l[2] = item_quantity
                    d_zagotovka[item_text] = l

                    f_zagotovka()                         # creating table 'selected items'
                    f_table_tree_inf()                    # creating MAIN table

                    if k_var == 11:
                        f_view_screen()                   # creating table SCREEN

                elif (tree_place_zapusk == tree_screen_rasschet) and k_var in [13]:      # ostatok CEH /tree_screen_rasschet
                    for key, value in d_ostatky.items():
                        if value[1] == item_text:
                            tmp = d_ostatky_copy_screen[key]
                            tmp[3] = item_quantity                                        # quantity_ceh
                            d_ostatky_copy_screen[key] = tmp
                            focus_item = list(filter(lambda x: x[1] == item_text, l_position))[0][0]  # position for FOCUS
                            break

                    f_rasschet(event, *args[0], focus_item=focus_item)


        except ValueError:
            showinfo('WRONG number', '{:^30}\n{:^30}\n{:^30}'.format('Only', 'NUMBERS', 'are alowed'))

        zapusk.destroy()  # close window = ORDER

# creating table 'selected items'
def f_table_middle():
    # deleting rows before laoding data
    rows = tree_middle.get_children()
    for item in rows:
        tree_middle.delete(item)

    # creating table 'selected items'
    counter = 1
    for keys, values in d_selected.items():
         nomenklatura = keys
         quantity = values
         tree_middle.insert('', counter, text=nomenklatura, values=(quantity))
         counter += 1

# deleting items in the 'table_middle'
def f_del_middle(event):
    item = tree_middle.item(tree_middle.selection())
    del d_selected[item['text']]
    f_table_middle() # creating table after removing item

def f_win_specif(tree_place):
    item = tree_place.item(tree_place.focus())  # item position
    item_text = item['text']

    # control presents in 'd_specification'
    if item_text in d_specification:
        win_specif = Toplevel(root, bg='lightblue',bd=5, relief=SUNKEN)
        win_specif.title('SPECIFICATION')
        win_specif.geometry('450x300')

        lab_specification = LabelFrame(win_specif, width=150, height=30, text=item_text, font='arial 12', relief=RAISED, bd=2)

        tree_specification = ttk.Treeview(win_specif)
        style = ttk.Style()
        style.configure('Treeview', font='arial 12')
        style.configure('Treeview.Heading', font='arial 12')

        tree_specification['column'] = ('quantity', 'place')
        tree_specification.column('quantity', width=100, anchor='n')
        tree_specification.column('place', width=100, anchor='n')
        tree_specification.heading('place', text='Участок')
        tree_specification.heading('quantity',text='к-во')

        lab_specification.grid(row=1, column=0,columnspan=1, padx=5, pady=5, sticky='w')
        tree_specification.grid(row=2, column=0,columnspan=1, padx=5, pady=5, sticky='w')

        d = d_specification.get(item_text, None)
        counter = 1
        for keys in sorted(d):
             tree_specification.insert('',counter, text=keys, values=(d.get(keys), d_karta_detaley.get(keys, None)[0]))
             counter += 1

        win_specif.focus()
        win_specif.bind('<Escape>', lambda x: win_specif.destroy())

    # if specification is not found
    else:
        showinfo('Ошибка', '{:^20}\n{:^20}\n{:^20}'.format('СПЕЦИФИКАЦИЯ для', item_text, 'ОТСУТСТВУЕТ'))



def f_zagotovka_zapusk():
    f_zagotovka_creation() # creation dictionary 'd_zagotovka'
    f_zagotovka()          # creation table 'tree_bottom'


def f_zagotovka_creation():
    global d_zagotovka, d_ostatky_kharkov, d_union_sel, d_karta_detaley, d_selected

    d_zagotovka = {}  # clearing dictionary befor writing
    # d_zagotovka {nomenklatura: [need, zapusk='', ostatky_kharkov, ostatky_ceh, delta_for_kharkov, shtribs, width, long, x, y, ostatok_nar]}
    d_length_shtribs = {}  # clearing dictionary befor writing
    # d_length_shtribs  {shtribs: quantity}

    for i in d_selected:
        # d_karta_detaley  {'Б 75 ЕР.Д': ['kharkov', '159', '67', '159', '14', '1']}
        # d_selected {'ЕР 16104/2В': 22}
        # d_specification {'ЕР 16104/2В': {'Б 75 ЕР.Д': 8, 'БС 164М.Боковая стенка.Д': 2, 'Втулка НМ 100_Элетон': 4, 'Дв 165/2ВМ.Дверь.Д': 2}
        if d_specification.get(i, None) == None:
            showinfo('Specification', '{:^30}\n{:^30}\n{:^30}'.format('Specification for', i, 'does not found'))
            # deleting rows before laoding data
            rows = tree_bottom.get_children()
            for item in rows:
                tree_bottom.delete(item)
            return None
    for keys, values in d_selected.items():
        if d_specification.get(keys, None) == None:
            print('{}\nSpecification is not found'.format(i))

        else:
            for d_sp_keys, d_sp_values in d_specification.get(keys).items():
                for values_ost in d_ostatky.values():
                    if values_ost[1] == d_sp_keys:
                        ostatky_ceh = values_ost[3]
                        break

                try:
                    ostatky_kharkov = '' if d_ostatky_kharkov.get(d_sp_keys,None) == None else int(d_ostatky_kharkov.get(d_sp_keys,None))
                except ValueError:
                    ostatky_kharkov = ''

                shtribs = d_karta_detaley.get(d_sp_keys, None)[3]
                width   = d_karta_detaley.get(d_sp_keys, None)[1]
                long    = d_karta_detaley.get(d_sp_keys, None)[2]
                x       = d_karta_detaley.get(d_sp_keys, None)[4]
                y       = d_karta_detaley.get(d_sp_keys, None)[5]

                # getting data for 'ostatok_nar'
                try:
                    naryad = 0
                    vipusk = 0
                    for data in d_union_sel.get(d_sp_keys, None):
                        naryad += data[0][3]
                        if data[1] == '':
                            vipusk += 0
                        else:
                            for ii in data[1]:
                                vipusk += ii[3]

                        ostatok_nar = int(naryad - vipusk)

                except TypeError:
                    ostatok_nar = 0


                if d_zagotovka.get(d_sp_keys, None) == None:
                    d_zagotovka[d_sp_keys] = [int(d_sp_values) * int(values), '', ostatky_kharkov, ostatky_ceh, '', shtribs, width, long, x, y, ostatok_nar] # new record
                else:
                    need_zag = d_zagotovka[d_sp_keys][0] + int(d_sp_values) * int(values)
                    d_zagotovka[d_sp_keys] = [need_zag, '', ostatky_kharkov, ostatky_ceh, '', shtribs, width, long, x, y, ostatok_nar] # editing previous record

def f_zagotovka_data(l):
    # [22, '', 1, 67, '', '159', '79', '390', '1', '2', 0]
    # [88, '', 0, 102, '', '159', '159', '898', '1', '1', 0]
    # [88, '', 0, 190, '', '138', '138', '593', '1', '1', 0]
    # [44, '', 134, 2, '', '138', '138', '1\xa0498', '1', '1', 0]
    # [88, '', 0, 13, '', '138', '138', '1\xa0600', '1', '1', 0]
    # [44, '', 0, 324, '', '138', '69', '130', '1', '2', 0]
    # [8, '', 6, 856, '', '159', '159', '67', '14', '1', 0]
    # [2, '', 0, '', '', '395', '395', '1\xa0560', '1', '1', 0]
    # [4, '', '', 176, '', '', '12', '39', '', '', 224]
    # [2, '', '', '', '', '', '522', '1\xa0582', '', '', 0]
    # [2, '', '', '', '', '', '456', '1\xa0530', '', '', 0]
    # [1, '', '', '', '', '', '1\xa0040', '470', '', '', 13]
    # [4, '', '', 1960, '', '', '100', '92', '', '', 0]

    # d_zagotovka {nomenklatura: [need, zapusk='', ostatky_kharkov, ostatky_ceh, delta_for_kharkov, shtribs, width, long, x, y, ostatok_nar}
    ostatok_ceh = 0 if l[3] == '' else int(l[3])
    ostatok_nar = l[10] if l[10] != 0 else '' # for view in cell table
    need = l[0]
    width = l[6]
    long = l[7]
    shtribs = l[5]
    x = l[8]
    y = l[9]
    # for ostatky_kharkov
    if l[2] == '':
        ostatky_kharkov = ''
        # delta = ostatok_ceh - need + l[10]  # l[10] - ostatok_nar  TEMPORARY
        delta = ostatok_ceh - need
    else:
        ostatky_kharkov = l[2]
        # delta = ostatky_kharkov - need + l[10]  # l[10] - ostatok_nar  TEMPORARY
        delta = ostatky_kharkov - need
    # for data, were zapusk was edited
    if l[1] == '':
        zapusk = abs(delta) if delta < 0 else ''
    else:
        zapusk = l[1]
    # for items where 'x' and 'y' don't exist
    if (x == '' or y == ''):
        zapusk_karta = long_all = delta_for_kharkov = ''
    else:
        x = int(x)
        y = int(y)

        zapusk = 0 if zapusk == '' else zapusk
        zapusk_karta = zapusk if zapusk % (x * y) == 0 else ((x * y) * (zapusk // (x * y)) + x * y)
        long_all = int(int(long.replace('\xa0', '')) * y * (zapusk_karta / x))
        # writing in 'd_zagotovka'[4] 'delta_for_kharkov' and d_zagotovka[i][1] 'zapusk'

        #delta_for_kharkov = '' if ((zapusk_karta + delta) == 0 or zapusk_karta==0) else (zapusk_karta + delta)
        delta_for_kharkov =  zapusk_karta + delta

        zapusk = '' if zapusk == 0 else zapusk  # for table view
        zapusk_karta = '' if zapusk_karta == 0 else zapusk_karta  # for table view
        long_all = '' if long_all == 0 else long_all  # for table view

    ostatok_ceh = '' if ostatok_ceh == 0 else ostatok_ceh  # '' instead '0'  put in table if ZERO

    return ostatok_ceh, ostatky_kharkov, need, delta, zapusk, width, long, shtribs, x, y, long_all, delta_for_kharkov, zapusk_karta, ostatok_nar

def f_zagotovka():
    global d_zagotovka, d_length_shtribs
    d_length_shtribs = {}

    # deleting rows before laoding data
    rows = tree_bottom.get_children()
    for item in rows:
        tree_bottom.delete(item)

    tree_bottom.tag_configure('Empty_store', background='lightblue', foreground='red')

    folder_all = tree_bottom.insert('', 1, text= 'Полуфабрикаты', values=())
    folder_kharkov = tree_bottom.insert('', 2, text='Харьковский', values=())
    folder_work = tree_bottom.insert('', 3, text='Заготовка', values=())
    folder_shtribs = tree_bottom.insert('', 4, text='Штрибс', values=())

    counter_all = 1
    counter_kharkov = 1
    counter_work = 1

    for i in sorted(d_zagotovka):
        # d_zagotovka {nomenklatura: [need, zapusk='', ostatky_kharkov, ostatky_ceh, delta_for_kharkov, shtribs, width, long, x, y]}
        ostatok_ceh, ostatky_kharkov, need, delta, zapusk, width, long, shtribs, x, y, long_all, delta_for_kharkov, zapusk_karta, ostatok_nar = f_zagotovka_data (d_zagotovka.get(i))

        # for 'folder_all'
        if delta < 0:
             tree_bottom.insert(folder_all, counter_all, text=i, values=(ostatok_ceh, ostatky_kharkov, ostatok_nar, need, delta, zapusk, width, long, shtribs, x, y, long_all, delta_for_kharkov, zapusk_karta), tags='Empty_store')
             counter_all += 1
        else:
             tree_bottom.insert(folder_all, counter_all, text=i, values=(ostatok_ceh, ostatky_kharkov, ostatok_nar, need, delta, zapusk, width, long, shtribs, x, y, long_all, delta_for_kharkov, zapusk_karta))
             counter_all += 1

        # for 'folder_kharkov'
        if d_karta_detaley.get(i)[0] == 'kharkov':
           if delta < 0:
               tree_bottom.insert(folder_kharkov, counter_kharkov, text=i, values=(ostatok_ceh, ostatky_kharkov, ostatok_nar, need, delta, zapusk, width, long, shtribs, x, y, long_all,delta_for_kharkov, zapusk_karta), tags='Empty_store')
               counter_kharkov += 1
           else:
               tree_bottom.insert(folder_kharkov, counter_kharkov, text=i, values=(ostatok_ceh, ostatky_kharkov, ostatok_nar, need, delta, zapusk, width, long, shtribs, x, y, long_all,delta_for_kharkov, zapusk_karta))
               counter_kharkov += 1

        # for 'folder_work'
        if d_karta_detaley.get(i)[0] == 'work':
           if delta < 0:
              tree_bottom.insert(folder_work, counter_work, text=i, values=(ostatok_ceh, ostatky_kharkov, ostatok_nar, need, delta, zapusk, width, long), tags='Empty_store')
              counter_work += 1
           else:
              tree_bottom.insert(folder_work, counter_work, text=i, values=(ostatok_ceh, ostatky_kharkov, ostatok_nar, need, delta, zapusk, width, long))
              counter_work += 1

    # creating list 'l_shtribs'
    counter_shtribs = 1
    counter = 1

    l_shtribs = []
    for i in d_zagotovka.values():
        if i[5] != '' and i[5] not in  l_shtribs:
            l_shtribs.append(i[5])

    # creating table 'folder shtribs'
    for i_shtr in l_shtribs:
        folder = tree_bottom.insert(folder_shtribs, counter_shtribs, text=i_shtr, values=())
        counter_shtribs += 1
        for i in sorted(d_zagotovka):
            if d_zagotovka.get(i)[5] == i_shtr:
                # d_zagotovka {nomenklatura: [need, zapusk='', ostatky_kharkov, ostatky_ceh, delta_for_kharkov, shtribs, width, long, x, y]}
                ostatok_ceh, ostatky_kharkov, need, delta, zapusk, width, long, shtribs, x, y, long_all, delta_for_kharkov, zapusk_karta, ostatok_nar = f_zagotovka_data(d_zagotovka.get(i))
                if delta <= 0:
                     tree_bottom.insert(folder, counter, text=i, values=(ostatok_ceh, ostatky_kharkov, ostatok_nar, need, delta, zapusk, width, long, shtribs, x, y, long_all, delta_for_kharkov, zapusk_karta),tags='Empty_store')
                else:
                     tree_bottom.insert(folder, counter, text=i, values=(ostatok_ceh, ostatky_kharkov, ostatok_nar, need, delta, zapusk, width, long, shtribs, x, y, long_all,delta_for_kharkov, zapusk_karta))
                counter += 1
                d_zagotovka[i][4] = d_zagotovka[i][4] if delta_for_kharkov == '' else delta_for_kharkov
                # adding data to 'd_length_shtribs'
                d_length_shtribs[i_shtr] = d_length_shtribs.get(i_shtr, 0) + (0 if long_all == '' else long_all)


    # deleting rows before laoding data
    rows = tree_shtribs.get_children()
    for item in rows:
        tree_shtribs.delete(item)

    # creating table 'shtribs'
    counter = 1
    for i in d_length_shtribs:
        l = d_length_shtribs.get(i)
        tree_shtribs.insert('', counter, text=i, values=(l/1000))  # convert from 'mm' to 'm'
        counter += 1


# Get curent date/time
def f_time_now():
    return datetime.now().strftime('%Y-%m-%d %H:%M:%S')

# saving information to 'product orders.txt'
def f_save_file():
    global d_length_shtribs
    # d_zagotovka {nomenklatura: [need, zapusk='', ostatky_kharkov, ostatky_ceh, delta_for_kharkov, shtribs, width, long, x, y]}

    if askyesno('Save file','{:^30}\n{:^30}'.format('Вы хотите сохранить', 'ЗАПУСК в файл?')):
        s = ''
        s += 'Запуск {}\n'.format(f_time_now())
        for i in d_selected:
            s += '{}    {}\n'.format(i, d_selected.get(i))
        s += 15* '-  '+'\n' # separator
        s += '{:^40}\n'.format('Харьковский станок')
        for i in sorted(d_zagotovka): # kharkov
            if d_karta_detaley.get(i)[0] == 'kharkov':
                ostatok_ceh, ostatky_kharkov, need, delta, zapusk, width, long, shtribs, x, y, long_all, delta_for_kharkov, zapusk_karta, ostatok_nar = f_zagotovka_data(d_zagotovka.get(i))
                s += '{:30.30} нужно: {:>4.4} ; запуск: {:>4.4} ; остаток с запуска: {:>4.4}\n\n'.format(i, str(need), str(zapusk), str(delta_for_kharkov))
        s += '{:^40}\n'.format('Заготовительный участок')
        for i in sorted(d_zagotovka): # work
            if d_karta_detaley.get(i)[0]  == 'work':
                ostatok_ceh, ostatky_kharkov, need, delta, zapusk, width, long, shtribs, x, y, long_all, delta_for_kharkov, zapusk_karta, ostatok_nar  = f_zagotovka_data(d_zagotovka.get(i))
                s += '{:30.30} нужно: {:>4.4} ; запуск: {:>4.4} ; остаток с запуска: {:>4.4}\n\n'.format(i, str(need), str(zapusk), str(delta_for_kharkov))
        s += 15 * '-  ' + '\n'  # separator
        for i in d_length_shtribs:
            s += 'Штрибс: {:>10}: Длина {:>10}\n'.format(i, d_length_shtribs.get(i)/1000)
        s += 15 * '-  ' + '\n'  # separator

        if not os.path.exists('product_orders.txt'):
            with open('product_orders.txt', 'w', encoding='utf-8') as f:
                f.write(s)
        else:
            with open('product_orders.txt', 'r+', encoding='utf-8') as f:
                data = f.read()
                s += data
                f.write(s)

        showinfo('ЗАПУСК', '{:^30}\n{:^30}'.format('Информация сохранена в', 'product_orders.txt'))

# saving by button DELTA
def f_save_delta():
    global d_ostatky_kharkov
    if askyesno('Save DELTA', '{:^40}\n{:^40}'.format('Вы хотите сохранить ОСТАТКИ запуска', 'для последующего использования')):

        for i in d_zagotovka:
            if d_zagotovka.get(i)[4] != '':
               d_ostatky_kharkov[i] = d_zagotovka.get(i)[4]

        d_ostatky_kharkov_new = {}  # removing from dictionary empty data
        for i in d_ostatky_kharkov:
            if d_ostatky_kharkov.get(i) != '':
                d_ostatky_kharkov_new[i] = d_ostatky_kharkov.get(i)
        d_ostatky_kharkov = d_ostatky_kharkov_new

        with open('ostatky_kharkov.pkl', 'wb') as f:
            pickle.dump(d_ostatky_kharkov, f)

        with open('ostatky_kharkov.txt', 'w', encoding='utf-8' ) as f:
            s = ''
            for keys, values in d_ostatky_kharkov.items():
                if values != '':
                   s += '{:40.40} :  {}\n'.format(keys, values)
            f.write(s)
        showinfo('DELTA', '{:^30}\n{:^30}'.format('Информация сохранена в', 'ostatky_kharkov.txt'))

# for loadinf data in 'd_ostatky_kharkov'
def f_load_delta():
    global d_ostatky_kharkov

    if not os.path.exists('ostatky_kharkov.pkl'):
        with open('ostatky_kharkov.pkl', 'wb') as f:
             pickle.dump({}, f)
             print("'ostatky_kharkov.pkl' was created")

    with open('ostatky_kharkov.pkl', 'rb') as f:
        data =  pickle.load(f)
        d_ostatky_kharkov = data

def f_delete_item(event):
    global d_zagotovka
    item_del = tree_bottom.item(tree_bottom.selection())
    if askyesno('Delete item', '{:^30}\n{:^30}\n{:^30}'.format('Do you really want to delete', item_del['text'], 'from table')):
       del d_zagotovka[item_del['text']]
       f_zagotovka()

# SCREEN
def f_open_screen():
    global   tree_screen, tree_screen_shtribs, fra_screen
    fra_screen = Toplevel()
    fra_screen.geometry('1800x900')
    fra_screen.title('SCREEN')

    rad_ostatok_screen = Radiobutton(fra_screen, text='Остаток', variable=var_screen_z, value=11)
    rad_zapusk_screen  = Radiobutton(fra_screen, text='Запуск', variable=var_screen_z, value=12)
    rad_product_screen = Radiobutton(fra_screen, text='Производство', variable=var_screen_z, value=16)
    btn_screen_view = MyButton(fra_screen, text='ПОКАЗАТЬ', command=f_view_screen)
    rad_all_screen = Radiobutton(fra_screen, text='Продукция', variable=var_screen, value=13)
    rad_kharkov_screen = Radiobutton(fra_screen, text='Харьковский', variable=var_screen, value=14)
    rad_zagotovka_screen = Radiobutton(fra_screen, text='Заготовка', variable=var_screen, value=15)
    btn_file = MyButton(fra_screen, text='Сохранить в файл', command=f_save_file)
    btn_delta = MyButton(fra_screen, text='Сохранить DELTA', command=f_save_delta)
    fra_screen_main = MyLabelFrame(fra_screen, width=1100, height=700, text='SCREEN', bg='bisque')
    fra_screen_shtribs = LabelFrame(fra_screen, width=300, height=300, text='Shtribs', bg='bisque')
    btn_screen_product = MyButton(fra_screen, text='ПРОИЗВОДСТВО', command=f_ostatky_product)

    rad_ostatok_screen.grid(row=1, column=0, padx=5, pady=5, sticky='w')
    rad_zapusk_screen.grid(row=1, column=1, padx=5, pady=5, sticky='w')
    rad_product_screen.grid(row=1, column=2, padx=5, pady=5, sticky='w')
    btn_screen_view.grid(row=1, column=3, padx=5, pady=5, sticky='w')
    rad_all_screen.grid(row=1, column=4, padx=5, pady=5, sticky='w')
    rad_kharkov_screen.grid(row=1, column=5, padx=5, pady=5, sticky='w')
    rad_zagotovka_screen.grid(row=1, column=6, padx=5, pady=5, sticky='w')
    btn_delta.grid(row=1, column=7, padx=5, pady=5, sticky='w')
    btn_file.grid(row=1, column=8, padx=5, pady=5, sticky='w')
    fra_screen_main.grid(row=2, column=0, columnspan=7)
    fra_screen_shtribs.grid(row=2, column=7, columnspan=8)
    btn_screen_product.grid(row=3, column=0, padx=5, pady=5, sticky='w')

    tree_screen = ttk.Treeview(fra_screen_main)
    tree_screen['height'] = 35
    tree_screen['column'] = (
        'ostatok_ceh', 'ostatok', 'Производство', 'need', 'delta', 'zapusk', 'width', 'long', 'shtribs', 'x', 'y', 'LONG', 'DELTA',
        'zapusk_karta')
    tree_screen.column('ostatok_ceh', width=80, anchor='n')
    tree_screen.column('ostatok', width=80, anchor='n')
    tree_screen.column('Производство', width=150, anchor='n')
    tree_screen.column('need', width=80, anchor='n')
    tree_screen.column('delta', width=80, anchor='n')
    tree_screen.column('zapusk', width=80, anchor='n')
    tree_screen.column('width', width=80, anchor='n')
    tree_screen.column('long', width=80, anchor='n')
    tree_screen.column('shtribs', width=80, anchor='n')
    tree_screen.column('x', width=80, anchor='n')
    tree_screen.column('y', width=80, anchor='n')
    tree_screen.column('LONG', width=100, anchor='n')
    tree_screen.column('DELTA', width=100, anchor='n')
    tree_screen.column('zapusk_karta', width=100, anchor='n')
    tree_screen.heading('ostatok_ceh', text='Цех')
    tree_screen.heading('ostatok', text='Остаток')
    tree_screen.heading('Производство', text='Производство')
    tree_screen.heading('need', text='НУЖНО')
    tree_screen.heading('delta', text='ДЕЛЬТА')
    tree_screen.heading('zapusk', text='ЗАПУСК')
    tree_screen.heading('width', text='ширина')
    tree_screen.heading('long', text='длина')
    tree_screen.heading('shtribs', text='штрибс')
    tree_screen.heading('x', text='x')
    tree_screen.heading('y', text='y')
    tree_screen.heading('LONG', text='ДЛИНА')
    tree_screen.heading('DELTA', text='DELTA')
    tree_screen.heading('zapusk_karta', text='Запуск_карта')
    tree_screen.grid(row=1, column=0)

    tree_screen_shtribs = ttk.Treeview(fra_screen_shtribs)
    tree_screen_shtribs['column'] = ('quantity')
    tree_screen_shtribs.column('quantity', width=100, anchor='n')
    tree_screen_shtribs.heading('quantity', text='Длина, м')
    tree_screen_shtribs.grid(row=1, column=0)

    tree_screen.bind('<Double-Button-1>', f_double_click_screen)
    tree_screen.bind('<Return>', f_double_click_screen)
    tree_screen.bind('<Delete>', f_delete_item_screen)

    f_view_screen()  # creating table SCREEN

# Bottom on screen
def f_view_screen():
    global d_zagotovka, d_length_shtribs, tree_screen, tree_screen_shtribs, fra_screen

    # deleting rows before laoding data
    rows = tree_screen.get_children()
    for item in rows:
        tree_screen.delete(item)

    tree_screen.tag_configure('Empty_store', background='lightblue', foreground='red')

    if var_screen.get() == 13:  # all production
        counter_all = 1
        # d_zagotovka {nomenklatura: [need, zapusk='', ostatky_kharkov, ostatky_ceh, delta_for_kharkov, shtribs, width, long, x, y]}
        for i in sorted(d_zagotovka):
            ostatok_ceh, ostatky_kharkov, need, delta, zapusk, width, long, shtribs, x, y, long_all, delta_for_kharkov, zapusk_karta, ostatok_nar = f_zagotovka_data (d_zagotovka.get(i))
            # for 'folder_all'
            if delta < 0:
                tree_screen.insert('', counter_all, text=i, values=(ostatok_ceh, ostatky_kharkov, ostatok_nar, need, delta, zapusk, width, long, shtribs, x, y, long_all, delta_for_kharkov, zapusk_karta), tags='Empty_store')
                counter_all += 1
            else:
                tree_screen.insert('', counter_all, text=i, values=(ostatok_ceh, ostatky_kharkov, ostatok_nar, need, delta, zapusk, width, long, shtribs, x, y, long_all, delta_for_kharkov, zapusk_karta))
                counter_all += 1

    elif var_screen.get() == 14:  # kharkov
        counter_kharkov = 1
        for i in sorted(d_zagotovka):
            ostatok_ceh, ostatky_kharkov, need, delta, zapusk, width, long, shtribs, x, y, long_all, delta_for_kharkov, zapusk_karta, ostatok_nar = f_zagotovka_data (d_zagotovka.get(i))
            # for 'folder_all'
            if d_karta_detaley.get(i)[0] == 'kharkov':
                if delta < 0:
                    tree_screen.insert('', counter_kharkov, text=i, values=(ostatok_ceh, ostatky_kharkov, ostatok_nar, need, delta, zapusk, width, long, shtribs, x, y, long_all,delta_for_kharkov, zapusk_karta), tags='Empty_store')
                    counter_kharkov += 1
                else:
                    tree_screen.insert('', counter_kharkov, text=i, values=(ostatok_ceh, ostatky_kharkov, ostatok_nar, need, delta, zapusk, width, long, shtribs, x, y, long_all,delta_for_kharkov, zapusk_karta))
                    counter_kharkov += 1

    elif var_screen.get() == 15:  # zagotovka
        counter_work = 1
        for i in sorted(d_zagotovka):
            ostatok_ceh, ostatky_kharkov, need, delta, zapusk, width, long, shtribs, x, y, long_all, delta_for_kharkov, zapusk_karta, ostatok_nar = f_zagotovka_data(d_zagotovka.get(i))
            if d_karta_detaley.get(i)[0] == 'work':
                 if delta < 0:
                       tree_screen.insert('', counter_work, text=i, values=(ostatok_ceh, ostatky_kharkov, ostatok_nar, need, delta, zapusk, width, long), tags='Empty_store')
                       counter_work += 1
                 else:
                       tree_screen.insert('', counter_work, text=i, values=(ostatok_ceh, ostatky_kharkov, ostatok_nar, need, delta, zapusk, width, long))
                       counter_work += 1

    # deleting rows before laoding data
    rows = tree_screen_shtribs.get_children()
    for item in rows:
        tree_screen_shtribs.delete(item)

    # creating table 'shtribs'
    counter = 1
    for i in d_length_shtribs:
        l = d_length_shtribs.get(i)
        tree_screen_shtribs.insert('', counter, text=i, values=(l/1000))  # convert from 'mm' to 'm'
        counter += 1

    #preventing from EMPTY table
    if d_zagotovka:
        item_selected = tree_screen.get_children()[0] # for lighting  the first position in the row
        tree_screen.selection_set(item_selected)
        tree_screen.focus_set()
        tree_screen.focus(item_selected)
    else:
        fra_screen.destroy()
        showinfo('ОШИБКА', 'Ничего НЕ ЗАПУЩЕНО!')


def f_delete_item_screen(event):
    global d_zagotovka
    item_del = tree_screen.item(tree_screen.selection())
    if askyesno('Delete item', 'Do you really want to delete\n{}\nfrom table'.format(item_del['text'])):
       del d_zagotovka[item_del['text']]
       f_zagotovka()  # creating table 'selected items'
       f_view_screen()  # creating table SCREEN

def f_save_store():
    global save_store, ent_save_store

    # window SAVE STORE
    save_store = Toplevel(root, bg='lightblue',bd=5, relief=SUNKEN)
    save_store.title('Saving selected')
    save_store.geometry('400x150')

    save_store_title = MyLabel(save_store, text='Введите название запуска')
    ent_save_store = Entry(save_store, width=40, bd=5, font='arial 12', fg='black', relief=SUNKEN)
    ent_save_store.focus()
    btn_save_store = MyButton(save_store,text='SAVE')
    save_store_title.grid(row=1, column=0, columnspan=3, padx=5, pady=5)
    ent_save_store.grid(row=2, column=0, padx=5, pady=5)
    btn_save_store.grid(row=3, column=0, padx=5, pady=5)

    btn_save_store.bind('<Button-1>', f_save_store_to)
    ent_save_store.bind('<Return>', f_save_store_to)

def f_save_store_to(event):
    global save_store, ent_save_store, d_saved_zapusk

    with open ('saved_zapusk.pkl', 'rb') as f:
        d_saved_zapusk = pickle.load(f)

    d_saved_zapusk[ent_save_store.get()] = [f_time_now(), d_selected]

    with open ('saved_zapusk.pkl', 'wb') as f:
        pickle.dump(d_saved_zapusk, f)

    previous_zapusk()
    save_store.destroy() # # close window = 'save_store'

def previous_zapusk():
    global d_saved_zapusk

    if not os.path.exists('saved_zapusk.pkl'):
        with open('saved_zapusk.pkl', 'wb') as f:
            pickle.dump(d_saved_zapusk, f)
            print("'saved_zapusk.pkl' was created")

    with open ('saved_zapusk.pkl', 'rb') as f:
        d_saved_zapusk = pickle.load(f)
    # deleting rows before laoding data
    rows = tree_store.get_children()
    for item in rows:
        tree_store.delete(item)
    # creating table
    counter = 1
    for i in d_saved_zapusk:
        tree_store.insert('', counter, text=i, values=(d_saved_zapusk[i][0]))

def f_input(event):
    global d_selected
    item = tree_store.item(tree_store.selection())
    d_selected = d_saved_zapusk.get(item['text'], None)[1]
    f_table_middle()

def f_delete_item_store(event):
    global d_saved_zapusk
    item_del = tree_store.item(tree_store.selection())
    if askyesno('Delete item', 'Do you really want to delete\n{}\nfrom table'.format(item_del['text'])):
       del d_saved_zapusk[item_del['text']]
       with open('saved_zapusk.pkl', 'wb') as f:
           pickle.dump(d_saved_zapusk, f)
       previous_zapusk()

# deleting items from 'screen_rasschet'  with DELETE
def f_delete_item_screen_rasschet(event, f1, f2, f3, f4, f5, f6):
    global d_ostatky_copy_screen, tree_screen_rasschet, l_position

    #l_position = [(0, 'Б 75 ЕР.Д'), (1, 'Вставка Опоры ВО 100.Д'), (2, 'Вставка Опоры.ВО 200.Д'), (3, 'Основание Опоры.ОО100.Д'),

    # d_ostatky_copy_screen = {00000000422 : ['Полуфабрикаты', 'ДВ 54М.Дверь.Д', '', 177, Decimal('20'), 1349, '', '0', '']}

    # item_del = {'text': 'УСЦ 400.Усилитель центральный ЕР.Д', 'image': '',
    #             'values': [159, 390, 79, 1, 2, '', 78, 78, 513, '0.15', '', ''], 'open': 0, 'tags': ''}
    item_del = tree_screen_rasschet.item(tree_screen_rasschet.selection())

    for key, value in d_ostatky_copy_screen.items():
        if value[1] == item_del['text']:
            del d_ostatky_copy_screen[key]
            focus_item = list(filter(lambda x: x[1] == item_del['text'] , l_position))[0][0]   #position 'deleting item'
            focus_item = focus_item if focus_item < len(l_position) - 1 else  -1   # for 'deleting' last row
            f_rasschet_func(f1, f2, f3, f4, f5, f6, focus_item=focus_item)
            break

def f_ostatky_kharkov():
    global  tree_kharkov, ostatky_kharkov, d_ostatky_kharkov
    # window OSTATKY KHARKOV
    ostatky_kharkov = Toplevel(root, bg='lightblue', bd=5, relief=SUNKEN)
    ostatky_kharkov.title('OSTATKY KHARKOV')
    ostatky_kharkov.geometry('350x450')

    tree_kharkov = ttk.Treeview(ostatky_kharkov)
    tree_kharkov['height'] = 20
    tree_kharkov['column'] = ('К-во')
    tree_kharkov.column('К-во', width=100, anchor='n')
    tree_kharkov.heading('К-во', text='К-во')
    tree_kharkov.grid(row=1, column=0)

    with open('ostatky_kharkov.pkl', 'rb') as f:
        d_ostatky_kharkov = pickle.load(f)

    counter = 1
    for i in sorted(d_ostatky_kharkov):
        tree_kharkov.insert('', counter, text=i, values=(d_ostatky_kharkov.get(i, None)))
        counter += 1
    tree_kharkov.bind('<Double-Button-1>', f_input_kharkov)
    tree_kharkov.bind('<Return>', f_input_kharkov)
    tree_kharkov.bind('<Delete>', f_delete_item_kharkov)

def f_delete_item_kharkov(event):
    global tree_kharkov, ostatky_kharkov
    item_del = tree_kharkov.item(tree_kharkov.selection())
    if askyesno('Delete item', 'Do you really want to delete\n{}\nfrom table'.format(item_del['text'])):
        del d_ostatky_kharkov[item_del['text']]
        with open('ostatky_kharkov.pkl', 'wb') as f:
             pickle.dump(d_ostatky_kharkov, f)
        ostatky_kharkov.destroy()
        f_ostatky_kharkov() # creating table  OSTATKY KHARKOV

def f_input_kharkov(event):
    global tree_kharkov, ent_kharkov_screen, kharkov_screen

    # for 'text' in 'ent_order'
    item_product = tree_kharkov.item(tree_kharkov.focus())
    item_text = item_product['text']
    item_quantity = item_product['values']

    # window PRODUCT
    kharkov_screen = Toplevel(root, bg='lightblue', bd=5, relief=SUNKEN)
    kharkov_screen.title('OSTATKY KHARKOV')
    kharkov_screen.geometry('400x150')
    lab_kharkov_screen_title = MyLabel(kharkov_screen, text='В запуск')
    lab_kharkov_screen_item = MyLabel(kharkov_screen, text=item_text)
    lab_kharkov_screen_quantity = MyLabel(kharkov_screen, text=item_quantity)
    ent_kharkov_screen = Entry(kharkov_screen, width=4, bd=5, font='arial 12', fg='black', relief=SUNKEN)
    btn_kharkov_screen = MyButton(kharkov_screen, text='SAVE')

    lab_kharkov_screen_title.grid(row=1, column=0, columnspan=3, padx=5, pady=5)
    lab_kharkov_screen_item.grid(row=2, column=0, columnspan=2, padx=5, pady=5)
    ent_kharkov_screen.grid(row=2, column=3, padx=5, pady=5)
    lab_kharkov_screen_quantity.grid(row=2, column=4, padx=5, pady=5)
    btn_kharkov_screen.grid(row=3, column=0, columnspan=4, padx=5, pady=5)

    btn_kharkov_screen.bind('<Button-1>', f_save_kharkov_screen)

def f_save_kharkov_screen(event):
    global d_ostatky_kharkov, tree_kharkov, ent_kharkov_screen, kharkov_screen

    item_product = tree_kharkov.item(tree_kharkov.focus())
    item_text = item_product['text']

    # control inputing number
    try:
        item_quantity = int(ent_kharkov_screen.get())
        if item_quantity >= 0:
            d_ostatky_kharkov[item_text] = item_quantity
            with open('ostatky_kharkov.pkl', 'wb') as f:
                pickle.dump(d_ostatky_kharkov, f)
        else:
            showinfo('WRONG number', '{:^25}\n{:^25}\n{:^25}'.format('Only', 'NUMBERS larger then ZERO', 'are alowed'))

    except ValueError:
        showinfo('WRONG number', '{:^25}\n{:^25}\n{:^25}'.format('Only', 'NUMBERS',  'are alowed(except ZERO)'))

    kharkov_screen.destroy()  # close window = ORDER
    ostatky_kharkov.destroy()
    f_ostatky_kharkov()  # creating table  OSTATKY KHARKOV

def f_connection():
    # 'flag_base_factory=False' for preventing rewriting 'base_factory.db' after WRONG connection to SQL-1C-BASE
    global data_naryad, data_vipusk, data_tovari, data_shtribs, data_zayavka, data_rezerv, data_czeny, flag_base_factory, data_karta_detaley
    print('Trying to Connect to SERVER................')
    try:
        server_my = helpic.server_my

        server = helpic.server
        database = helpic.database
        username = helpic.username
        password = helpic.password
        driver = helpic.driver          # Driver you need to connect to the database
        port = helpic.port
        # print(server, database, username, password, driver, port)
        # for testing on MY NOTEBOOK
        # cnxn = pyodbc.connect(
        #     'DRIVER=' + driver + ';SERVER=' + server_my + ';DATABASE=' + database + ';Trusted_Connection=yes;')

        # for working on SERVER
        cnxn = pyodbc.connect(
            'DRIVER=' + driver + ';PORT=port;SERVER=' + server + ';PORT=' + port + ';DATABASE=' + database + ';UID=' + username +
            ';PWD=' + password)
        # print('cnxn = ', cnxn)
        cursor = cnxn.cursor()
        print('Connection to SERVER is OK................')

        # _Reference50  - kontragent
        # _Reference64  - nomenklatura
        # _Reference75  - polzovateli
        # _Document236  - Vipusk produkcii
        # _Document232  - Naryadi
        # _Reference82  - Scladi
        # _Reference10250 - SHTRIBSI
        # _Reference9542  - Zayavki pokupateley
        # _AccumRg9658 - / OSNOVNAYA / Registr Nakopleniya Rezerv Realizacii
        # InfoRg8868 - Registr Svedeniy Czeny Nomenklaturi

        # _Reference64  - nomenklatura
        # [_IDRRef]
        # , [_Version]
        # , [_Marked]
        # , [_IsMetadata]
        # , [_ParentIDRRef]
        # , [_Folder]
        # , [_Code]                Code Nomenclatura
        # , [_Description]         Description Nomenklatura
        # , [_Fld1035]
        # , [_Fld1036RRef]
        # , [_Fld1037RRef]
        # , [_Fld1038]
        # , [_Fld1039]
        # , [_Fld1040RRef]
        # , [_Fld1041]
        # , [_Fld1042]
        # , [_Fld1043]
        # , [_Fld1044]
        # , [_Fld1045]
        # , [_Fld1046]
        # , [_Fld1047RRef]
        # , [_Fld1048RRef]
        # , [_Fld1049]
        # , [_Fld1050RRef]
        # , [_Fld1051_TYPE]
        # , [_Fld1051_RTRef]
        # , [_Fld1051_RRRef]
        # , [_Fld9531RRef]
        # , [_Fld10329]
        # , [_Fld10330RRef]
        # , [_Fld10485]

        # _AccumRg9285 - /OSNOVNAYA/ Registr Nakopleniya Tovari Na Skladah
        #    [_Period]         data
        # , [_RecorderTRef]    +  '000234'   /   -  '000236'  bytes hex
        # , [_RecorderRRef]      registrator /
        # , [_LineNo]
        # , [_Active]
        # , [_RecordKind]
        # , [_Fld9286RRef]      scladi       / _Reference82
        # , [_Fld9287RRef]      nomenklatura / _Reference64
        # , [_Fld9288RRef]
        # , [_Fld9289RRef]
        # , [_Fld9290RRef]
        # , [_Fld9291]          quantity
        # , [_Fld9292RRef]

        # AccumRg9681  - /OSNOVNAYA/ Registr Nakopleniya Dvigeniya Nomenklaturi Eleton
        # [_Period]
        # , [_RecorderTRef]       registrator document
        # , [_RecorderRRef]       registrator nomer documenta
        # , [_LineNo]
        # , [_Active]
        # , [_RecordKind]         +/-     '+' = 0,  '-' = 1
        # , [_Fld9682_TYPE]
        # , [_Fld9682_S]
        # , [_Fld9682_RTRef]
        # , [_Fld9682_RRRef]
        # , [_Fld9683RRef]          nomenklatura / _Reference64
        # , [_Fld9684RRef]          scladi       / _Reference82
        # , [_Fld9685]              quantity

        # _Reference10250  - Dictionary   SHTRIBSI
        #   [_IDRRef]
        # , [_Version]
        # , [_Marked]        -  DELETING mark
        # , [_IsMetadata]
        # , [_Code]
        # , [_Description]   - naimenovanie
        # , [_Fld10313]      - width
        # , [_Fld10252]      - length
        # , [_Fld10253]      - thickness
        # , [_Fld10254RRef]

        # _Reference9542  - Zayavki pokupateley
        # [_IDRRef]
        # , [_Version]
        # , [_Marked]
        # , [_IsMetadata]
        # , [_Code]
        # , [_Fld9543]             - date
        # , [_Fld9544RRef]         - kontragent
        # , [_Fld9545_TYPE]
        # , [_Fld9545_RTRef]       - document
        # , [_Fld9545_RRRef]
        # , [_Fld9546RRef]        - nomenklatura
        # , [_Fld9547]            - quantity
        # , [_Fld9548]
        # , [_Fld9549]
        # , [_Fld9550]            - primechanie
        # , [_Fld9551RRef]
        # , [_Fld9552RRef]
        # , [_Fld9553]
        # , [_Fld9554]
        # , [_Fld9555]
        # , [_Fld9556]
        # , [_Fld9557RRef]
        # , [_Fld9558RRef]
        # , [_Fld10872]

        # _AccumRg9658           -  /OSNOVNAYA/ Registr Nakopleniya Rezerv Realizacii
        # [_Period]              -  date
        # , [_RecorderTRef]
        # , [_RecorderRRef]
        # , [_LineNo]
        # , [_Active]
        # , [_RecordKind]
        # , [_Fld9659RRef]
        # , [_Fld9660RRef]       - nomenklatura
        # , [_Fld9661RRef]
        # , [_Fld9662RRef]       - kontragent
        # , [_Fld9663]           - quantity
        # , [_Fld9664]

        # InfoRg8868 - Registr Svedeniy Czeny Nomenklaturi
        # [_Period]              - date
        # , [_RecorderTRef]
        # , [_RecorderRRef]
        # , [_LineNo]
        # , [_Active]
        # , [_Fld8869RRef]
        # , [_Fld8870RRef]       - nomenklatura
        # , [_Fld8871RRef]
        # , [_Fld8872]

        # _Reference103  typi czen nomenklaturi


        #  _Reference9689 - karta detaley
        #     [_IDRRef]
        #     , [_Version]
        #     , [_Marked]
        #     , [_IsMetadata]
        #     , [_Code]
        #     , [_Fld9692RRef]   - nomenklatura
        #     , [_Fld9693]
        #     , [_Fld9694]       - width
        #     , [_Fld9695]
        #     , [_Fld9696]       - length
        #     , [_Fld9697RRef]   - koefficient othoda
        #     , [_Fld9698]
        #     , [_Fld9699RRef]
        #     , [_Fld9700]
        #     , [_Fld10056RRef]
        #     , [_Fld10058]
        #     , [_Fld10060]
        #     , [_Fld10246RRef]
        #     , [_Fld10066]
        #     , [_Fld10247]      - X
        #     , [_Fld10248]      - Y
        #     , [_Fld10451]
        #     , [_Fld10486]      - shtribs_1


        # !!!! I can not create 'ekran' for SQL QUARY  '?' - does not work  !!!

        quary_n = \
        'SELECT    [_Date_Time], [_Number], [_Marked]\
        ,(SELECT [_Description] FROM .[dbo].[_Reference50] WHERE [_IDRRef] = [_Fld7115RRef] ) \
        ,(SELECT [_Description] FROM .[dbo].[_Reference64] WHERE [_IDRRef] = [_Fld7126RRef] ) \
        ,[_Fld7127]\
        FROM [{}].[dbo].[_Document232]'.format(database)
        start = monotonic()
        data_naryad = cursor.execute(quary_n)
        data_naryad = data_naryad.fetchall()
        print("\033[0m information for 'data_naryad' is recieved. \033[32m {:.5f} seconds \033[0m ".format(monotonic()-start))

        quary_v = \
        'SELECT    [_Date_Time], [_Number], [_Marked]\
        ,(SELECT [_Description] FROM .[dbo].[_Reference50] WHERE [_IDRRef] = [_Fld7358RRef] ) \
        ,(SELECT [_Description] FROM .[dbo].[_Reference64] WHERE [_IDRRef] = [_Fld7346RRef] ) \
        ,[_Fld7354]\
        ,(SELECT [_Number] FROM .[dbo].[_Document232] WHERE [_IDRRef] = [_Fld7350RRef])\
        ,(SELECT [_Date_Time] FROM .[dbo].[_Document232] WHERE [_IDRRef] = [_Fld7350RRef])\
        FROM [{}].[dbo].[_Document236]'.format(database)
        start = monotonic()
        data_vipusk = cursor.execute(quary_v)
        data_vipusk = data_vipusk.fetchall()

        print("\033[0m information for 'data_vipusk' is recieved. \033[32m {:.5f} seconds \033[0m ".format(monotonic()-start))

        quary_tovari = 'SELECT  [_Period], [_RecordKind], [_Fld9685]\
      ,(SELECT[_Code] FROM.[dbo].[_Reference64] WHERE[_IDRRef] =[_Fld9683RRef]) \
      ,(SELECT [_Description] FROM .[dbo].[_Reference64] WHERE [_IDRRef] = [_Fld9683RRef] ) \
      ,(SELECT [_Description] FROM .[dbo].[_Reference82] WHERE [_IDRRef] = [_Fld9684RRef] ) \
      , [_Fld9682_RTRef]\
        FROM [{}].[dbo].[_AccumRg9681]'.format(database)
        start = monotonic()
        data_tovari = cursor.execute(quary_tovari)
        data_tovari = data_tovari.fetchall()
        print("\033[0m information for 'data_tovari' is recieved. \033[32m {:.5f} seconds \033[0m ".format(monotonic()-start))

        quary_shtribs = 'SELECT  * FROM [{}].[dbo].[_Reference10250]'.format(database)
        start = monotonic()
        data_shtribs = cursor.execute(quary_shtribs)
        data_shtribs = data_shtribs.fetchall()
        print("\033[0m information for 'data_shtribs' is recieved. \033[32m {:.5f} seconds \033[0m ".format(monotonic()-start))

        quary_zayavka = 'SELECT  [_Fld9543] \
    , (SELECT[_Description] FROM.[dbo].[_Reference50] WHERE[_IDRRef] =[_Fld9544RRef]) \
    , (SELECT[_Description] FROM.[dbo].[_Reference64] WHERE[_IDRRef] =[_Fld9546RRef]) \
    , [_Fld9547] ,[_Fld9550]\
    , (SELECT [_Code] FROM.[dbo].[_Reference64] WHERE[_IDRRef] =[_Fld9546RRef]) \
    FROM [{}].[dbo].[_Reference9542]'.format(database)
        start = monotonic()
        # data_zayavka = (datetime.datetime(4020, 1, 21, 8, 31, 56), 'УКРЕНЕРГО-АЛЬЯНС', 'ЕР 1864/1', Decimal('4.000'), '20.01.2020', '00000011693')
        data_zayavka = cursor.execute(quary_zayavka)
        data_zayavka = data_zayavka.fetchall()

        print("\033[0m information for 'data_zayavka' is recieved. \033[32m {:.5f} seconds \033[0m ".format(monotonic()-start))

        quary_rezerv = 'SELECT [_Period] \
    , (SELECT[_Description] FROM.[dbo].[_Reference50] WHERE[_IDRRef] =[_Fld9662RRef]) \
    , (SELECT[_Description] FROM.[dbo].[_Reference64] WHERE[_IDRRef] =[_Fld9660RRef]) \
    , [_Fld9663] \
    , (SELECT [_Code] FROM.[dbo].[_Reference64] WHERE[_IDRRef] =[_Fld9660RRef])\
    FROM [{}].[dbo].[_AccumRg9658]'.format(database)
        start = monotonic()
        data_rezerv = cursor.execute(quary_rezerv)
        data_rezerv = data_rezerv.fetchall()
        print("\033[0m information for 'data_rezerv' is recieved. \033[32m {:.5f} seconds \033[0m ".format(monotonic()-start))

        quary_czeny = 'SELECT [_Period],  [_Fld8872]\
      ,(SELECT[_Description] FROM.[dbo].[_Reference103] WHERE[_IDRRef] = [_Fld8869RRef])\
      ,(SELECT[_Code] FROM.[dbo].[_Reference64] WHERE[_IDRRef] = [_Fld8870RRef])\
      ,(SELECT[_Description] FROM.[dbo].[_Reference64] WHERE[_IDRRef] = [_Fld8870RRef])\
    FROM [{}].[dbo].[_InfoRg8868]'.format(database)
        start = monotonic()
        # data_czeny = (datetime.datetime(4019, 11, 18, 0, 0), Decimal('243.12'), 'РозничнаяП2', '00000014861', 'Ячейка К2 (ККУ)')
        data_czeny = cursor.execute(quary_czeny)
        data_czeny = data_czeny.fetchall()
        print("\033[0m information for 'data_czeny' is recieved. \033[32m {:.5f} seconds \033[0m ".format(monotonic() - start))

        quary_karta_detaley = 'SELECT [_Fld9694], [_Fld9696], [_Fld10486], [_Fld10247], [_Fld10248]\
        ,(SELECT[_Code] FROM.[dbo].[_Reference9690] WHERE[_IDRRef] = [_Fld9697RRef])\
        ,(SELECT[_Code] FROM.[dbo].[_Reference64] WHERE[_IDRRef] =[_Fld9692RRef])\
        ,(SELECT[_Description] FROM.[dbo].[_Reference64] WHERE[_IDRRef] = [_Fld9692RRef])\
         FROM [{}].[dbo].[_Reference9689]'.format(database)

        start = monotonic()
        # (Decimal('146.00'), Decimal('188.00'), Decimal('0'), Decimal('0'), Decimal('0'), 'рб                  ', ' Пластина боковая ПлБ41ТС.Д')
        data_karta_detaley = cursor.execute(quary_karta_detaley)
        data_karta_detaley = data_karta_detaley.fetchall()

        print("\033[0m information for 'data_karta_detaley' is recieved. \033[32m {:.5f} seconds \033[0m ".format(monotonic() - start))

        cnxn.close()
        flag_base_factory = True     # 'flag_base_factory=False' for preventing rewriting 'base_factory.db' after WRONG connection to SQL-1C-BASE
        print('Connection to SERVER is CLOSED ................')

    except:
        print('Can not connect ...............')
        flag_base_factory = False  # 'flag_base_factory=False' for preventing rewriting 'base_factory.db' after WRONG connection to SQL-1C-BASE
        #creating empty lists for TABLES
        data_naryad = []
        data_vipusk = []
        data_tovari = []
        data_shtribs = []
        data_zayavka = []
        data_rezerv = []


def f_nezavershonnie_naryadi():
    # NARYADI
    global d_union_sel, d_vipusk, d_naryad, d_original

    d_union = {}
    start = monotonic()
    # data_naryad
    # (datetime.datetime(4016, 10, 19, 8, 11, 32), '00000011291', b'\x00', 'Галиченерго', 'МКН 33.15М IP54', Decimal('5'))
    lst = list([i[0] for i in d_original.values()]) # list with nomenklatura
    # print('l', l)
    for i in data_naryad:
        # print(i[4], i[4] in l)
        # sleep(1)
        # if (i[4] in l) and (ord(i[2]) == 0):
            # print(i[4], i[4] in l)

        if ord(i[2]) == 0:
            if d_naryad.get(i[4], None) == None:
                d_naryad[i[4]] = [[i[0], i[1], i[3], i[5]]]  # data, nomer, kontragent, quantity
            else:
                l = d_naryad.get(i[4], None) + [[i[0], i[1], i[3], i[5]]]
                d_naryad[i[4]] = l

    # d_naryad  =  {nomenklatura: [date, number, quantity], ...., [...]}

    # control d_naryad
    # for i in d_naryad:
    #     print(i, d_naryad.get(i))

    print("\033[0m 'd_naryad' is completed \033[32m {:.5f} seconds \033[0m ".format(monotonic()-start))
    start = monotonic()
    for i in data_vipusk:
        if ord(i[2]) == 0:
            if d_vipusk.get(i[4], None) == None:
                d_vipusk[i[4]] = [[i[0], i[1], i[3], i[5], i[6],
                                   i[7]]]  # data, nomer, kontragent, quantity, nomer_naryada, data_naryada
            else:
                l = d_vipusk.get(i[4], None) + [[i[0], i[1], i[3], i[5], i[6], i[7]]]
                d_vipusk[i[4]] = l
    # d_vipusk  =  {nomenklatura: [[date, number, quantity, number naryad, date naryad], ...., [...]}

    # # control d_vipusk
    # for i in d_vipusk:
    #     print(i, d_vipusk.get(i))
    counter = 1
    print("\033[0m 'd_vipusk' is completed \033[32m {:.5f} seconds \033[0m ".format(monotonic()-start))
    print("'d_union_sel' - NE ZAVERSHONNIE  NARYADI - creating ................. ")
    start = monotonic()
    # start_1 = monotonic()
    lst = list([i[0] for i in d_original.values()])  # list with nomenklatura
    for nomenklatura in d_naryad:
        if   nomenklatura in lst:
            for narydi in d_naryad.get(nomenklatura):
                 try:
                    vip_nar = []
                    for i in d_vipusk.get(nomenklatura, None):
                        vip_nar = [i for i in d_vipusk.get(nomenklatura, None) if (i[4] == narydi[1] and i[5] == narydi[0])]
                 except TypeError:
                     vip_nar = []
            # if d_union.get(nomenklatura, None) == None:
                 if  nomenklatura not in d_union:
                    d_union[nomenklatura] = [[narydi, vip_nar]]  # [0] = [[narydi],[[vip_nar], [...], ... , [...]]]
                 else:
                # l = d_union.get(nomenklatura, None)
                # d_union[nomenklatura] = l + [[narydi, vip_nar]]
                      d_union[nomenklatura] += [[narydi, vip_nar]]
                # d_union[nomenklatura] = list (chain(l, [[narydi, vip_nar]]))
    # print(1, monotonic() - start_1)
    # start_2 = monotonic()
    for i in sorted(d_union):  # 'i' = NOMENKLATURA
        for data in d_union.get(i, None):
            nar = data[0]
            vip = data[1]
            sum = 0            # summa all vipusks in naryad
            for ii in vip:
                sum += ii[3]
            if sum < nar[3]:
                # if d_union_sel.get(i, None) == None:
                if i not in d_union_sel:
                    d_union_sel[i] = [[nar+[sum], vip]]
                else:
                    # l = d_union_sel.get(i, None)
                    d_union_sel[i] +=  [[nar+[sum], vip]] # summa all vipusks in naryad
    # print(2, monotonic() - start_2)
    print("\033[0m 'd_union_sel' is completed......... \033[32m {:.5f} seconds \033[0m ".format(monotonic() - start))

def f_ostatky_product():

    l_for_print = []
    counter = 1
    for data in d_union_sel:  # 'data' = NOMENKLATURA
        for i in d_union_sel[data]:
            l_for_print.append([data]+i[0])

    l_for = sorted(l_for_print, key=lambda x: x[1])

    if askyesno('Save ostatky_product','{:^25}\n{:^25}\n{:^25}'.format('Вы хотите сохранить', 'НЕ ЗАКРЫТЫЕ НАРЯДЫ', 'в файл?')):
        s = ''
        s += 'НЕ ЗАКРЫТЫЕ НАРЯДЫ на дату: {}\n'.format(f_time_now())
        counter = 1
        for i in l_for:

            item = i[0]
            date = i[1]
            numb_nar = i[2]
            kontr = i[3] if i[3] else ''
            quont = i[4]
            ost = i[4]-i[5]

            s += '{}; {:25.25};  {}; {:>12.12}; {:>20.20}; {:>4.4}; В ПРОИЗВОДСТВЕ: {:>4.4};\n\n'.\
                format(counter, item, date, numb_nar, kontr, str(quont), str(ost))
            counter += 1

        with open('narydi_ne_zakritie.txt', 'w', encoding='utf-8') as f:
             f.write(s)

        showinfo('НАРЯДЫ В ПРОИЗВОДСТВЕ', '{:^25}\n{:^25}'.format('Информация сохранена в' , 'narydi_ne_zakritie.txt') )

def f_ostatky_ceh_mk():
    global data_tovari, d_ostatky, d_original

    d_ostatky_full = {}
    print("Starting to create 'd_ostatky' for 'ceh' and 'mk'............")
    # // example //  'i' in 'data_tovari'
    # (datetime.datetime(4013, 6, 7, 12, 49, 3),  Decimal('1'), '00000009705', 'Усилитель УСЗС 18', 'Склад Цех №2')
    start = monotonic()
    # d_original = {'00000013791': ['ЕР 16104/2В', 'ЕР '],                   }
    # lst = ['00000012591',  ....... ]
    lst = list([i for i in d_original])  # list with nomenklatura

    # data_tovari = [
    # (datetime(4020, 1, 10, 15, 19, 39), Decimal('1'), Decimal(20.000'), '00000015320', 'СМГ 16.35М (1,5)', 'Склад М/К', b'\x00\x00\x00\xd9'),
    # (datetime.datetime(4020, 1, 10, 15, 19, 39), Decimal('1'), Decimal('20.000'), '00000012591', 'М 6.150 В', 'Склад М/К', b'\x00\x00\x00\xd9'),
    # (datetime.datetime(4020, 1, 10, 15, 19, 40), Decimal('1'), Decimal('1.000'),  '00000008579', 'МКН 663М IP54','Склад М/К', b'\x00\x00\x00\xd9')]
    for i in data_tovari:

        if i[3] in lst:
            p_m = 'PLUS' if i[1] == 0 else 'MINUS'
            # {Code : [nomenklatura, date, PLUS/MINUS, Sklad, Quantity], [...], ....  ,[....]}
            if d_ostatky_full.get(i[3], None) == None:
                d_ostatky_full[i[3]] = [[i[4], i[0], p_m, i[5], i[2]]]
            else:
                l = d_ostatky_full[i[3]]
                d_ostatky_full[i[3]] = l + [[i[4], i[0], p_m, i[5], i[2]]]

    # checking 'd_ostatky_full'
    # print("..... checking 'd_ostatky_full'......")
    # print('d_ostatky_full = ', d_ostatky_full)
    # for i in d_ostatky_full:
    #     print(i, d_ostatky_full.get(i, None) )

    # // example //  'i' in 'd_ostatky_full'
    # 00000016928[['Опора.О200.У Сборка_Окр', datetime.datetime(4019, 6, 18, 11, 24, 53), 'MINUS', 'Склад Цех №2',Decimal('16.000')],\
    #             ['Опора.О200.У Сборка_Окр', datetime.datetime(4019, 6, 18, 11, 27, 12), 'PLUS',  'Склад Цех №2', Decimal('16.000')]]

    # d_ostatky {Code: [group, nomenklatura, quantity_mk, quantity_ceh} - THE GOAL

    for i in d_original:
        code = i
        try:
            group = d_original[i][1]
            nomenklatura = d_original[i][0]
            try:
                plus_ceh = sum([ii[4] for ii in d_ostatky_full[i] if (ii[3] == 'Склад Цех №2' and ii[2] == 'PLUS')])
            except KeyError:
                plus_ceh = 0

            try:
                minus_ceh = sum([ii[4] for ii in d_ostatky_full[i] if (ii[3] == 'Склад Цех №2' and ii[2] == 'MINUS')])
            except KeyError:
                minus_ceh = 0
            quantity_ceh = int(plus_ceh) - int(minus_ceh) if (int(plus_ceh) - int(minus_ceh)) != 0 else ''

            try:
                plus_mk = sum([ii[4] for ii in d_ostatky_full[i] if (ii[3] == 'Склад М/К' and ii[2] == 'PLUS')])
            except KeyError:
                plus_mk = 0

            try:
                minus_mk = sum([ii[4] for ii in d_ostatky_full[i] if (ii[3] == 'Склад М/К' and ii[2] == 'MINUS')])
            except KeyError:
                minus_mk = 0
            quantity_mk = int(plus_mk) - int(minus_mk) if (int(plus_mk) - int(minus_mk)) != 0 else ''

            d_ostatky[i] = [group, nomenklatura, quantity_mk, quantity_ceh]

        except KeyError:
            print("'Exception' with {}".format(d_original[i][0]))

    print("\033[0m 'd_ostatky' is created\033[32m{: .5f} seconds \033[0m".format(monotonic()-start))

def f_refresh():
    global d_ostatky, d_naryad, d_vipusk, d_union_sel, text_refresh
    d_ostatky = {}

    d_naryad = {}     # clearing OLD dictionary
    d_vipusk = {}     # clearing OLD dictionary
    d_union_sel = {}  # clearing OLD dictionary

    f_connection()    # SQL-quary to 1C
    f_main()          # for 'WORK' and 'DEMO' versions

    f_load_delta()    # 'ostatky_kharkov.pkl' - load data
    f_table_tree_inf()# creating MAIN TABLE // 'table_tree_inf'
    previous_zapusk() # for loading data in table 'previous zapusk'
    f_create_db()     # rewriting 'base_factory.db' and sending to FTP-SERVER
    f_statistica()    # rewriting 'statistic.xls' from FTP-SERVER

    text_refresh = 'SYNC = ' + f_time_now()# date/time last syncronization with 1Cserver
    btn_refresh.config(text=text_refresh)  # date/time last syncronization with 1Cserver

    print('Syncronization with 1C is finished')

def f_shtribs():
    global data_shtribs

    l_data_shtribs = []
    for i in data_shtribs:
        if i[2] == b'\x00':  # position does not marked 'DEL'
            name = i[5]
            lenght = i[6]
            width = i[7]
            thickness = i[8]
            l_data_shtribs.append([name, lenght, width, thickness])

    l_data_shtribs.sort(key=lambda x: x[0] )

    win_shtribs = Toplevel(root, bg='lightblue', bd=5, relief=SUNKEN)
    win_shtribs.title('SHTRIBS')
    win_shtribs.geometry('500x500')

    tree_shtribs = ttk.Treeview(win_shtribs)
    tree_shtribs['height'] = 20
    tree_shtribs['column'] = ('width', 'long',  'thickness')
    tree_shtribs.column('long', width=100, anchor='n')
    tree_shtribs.column('width', width=100, anchor='n')
    tree_shtribs.column('thickness', width=100, anchor='n')

    tree_shtribs.heading('width', text='Ширина')
    tree_shtribs.heading('long', text='Длина')
    tree_shtribs.heading('thickness', text='Толщина')

    tree_shtribs.grid(row=1, column=0, columnspan=1, padx=5, pady=5, sticky='w')

    counter = 1
    for i in l_data_shtribs:
        tree_shtribs.insert('', counter, text=i[0], values=(i[1], i[2], i[3]))
        counter += 1

def f_zayavka():
    global d_zayavka, data_zayavka
    d_zayavka = {}

    for i in data_zayavka:
        # (datetime.datetime(4019, 9, 2, 8, 17, 44), 'СФЕРАЛАЙН ЛТД', 'МКН 442М IP31', Decimal('15.000'), '03.09.2019', '00000000574')
        if d_zayavka.get(i[5], None) == None:
            d_zayavka[i[5]] = [[i[0], i[1], i[3], i[4], i[2]]]
        else:
            l = d_zayavka.get(i[5], None)
            d_zayavka[i[5]] = l + [[i[0], i[1], i[3], i[4], i[2]]]
    # 'd_zayavka'
    # 00000012162 [[datetime.datetime(4019, 9, 4, 9, 25, 49), 'ЕТР-КОМПЛЕКС', Decimal('2.000'), '03.09.2019', 'М 4.75 В '],
    # [datetime.datetime(4019, 9, 3, 15, 13, 2), 'КМ-РІШЕННЯ', Decimal('8.000'), '03.09.2019', 'М 4.75 В '],
    # [datetime.datetime(4019, 9, 5, 15, 31, 43), 'Мадек', Decimal('20.000'), '', 'М 4.75 В '],
    # [datetime.datetime(4019, 8, 29, 13, 57, 31), 'СТИКС-ОИЛ ИНЖИНИРИНГ', Decimal('4.000'), '29.08.2019', 'М 4.75 В ']]

    # for checking 'd_zayavka'
    # for i in d_zayavka:
    #     print(i, d_zayavka[i])

def f_rezerv():
    global d_rezerv, data_rezerv
    d_rezerv = {}
    # 'data_rezerv'
    # (datetime.datetime(4019, 9, 2, 8, 20), 'Елетон Електрощитовий завод', 'МКН 1283М IP31', Decimal('1.00'), '00000007821')
    for i in data_rezerv:
        if d_rezerv.get(i[4], None) == None:
            d_rezerv[i[4]] = [[i[0], i[1], i[3], i[2]]]
        else:
            l = d_rezerv.get(i[4], None)
            d_rezerv[i[4]] = l + [[i[0], i[1], i[3], i[2]]]
    # 'd_rezerv'
    # 00000015596[[datetime.datetime(4019, 9, 4, 9, 1, 21), 'Промавтоматика-Вінниця', Decimal('4.00'), 'СМГ 18.35 (1,5)'],
    #             [datetime.datetime(4019, 9, 4, 15, 39, 11), 'СОЮЗ-СВІТЛО УКРАЇНА', Decimal('6.00'), 'СМГ 18.35 (1,5)'],
    #             [datetime.datetime(4019, 9, 2, 8, 20), 'ТЕХНОТОН ЕНЕРГО', Decimal('38.00'), 'СМГ 18.35 (1,5)'],
    #             [datetime.datetime(4019, 9, 5, 16, 25, 36), 'Українське електрообладнання та інсталяція', Decimal('2.00'), 'СМГ 18.35 (1,5)']]

    # for checking 'd_rezerv'
    # for i in d_rezerv:
    #     print(i, d_rezerv[i])

def f_report():   # start and end work, dinner time
    engine = pyttsx3.init()
    date_time_now = datetime.now().strftime("%m/%d/%Y, %H:%M")          # date_time_now
    time_now = datetime.now().strftime("%H:%M")                         # time_now
    date_object_time_now = datetime.strptime(time_now, "%H:%M")         # date_object_time_now

    time_dinner_string = '12:00'
    time_object_dinner = datetime.strptime(time_dinner_string, "%H:%M")  # date_object_dinner

    time_dinner_end_string = '13:00'
    time_object_dinner_end = datetime.strptime(time_dinner_end_string, "%H:%M")  # date_object_dinner

    time_end_work_string = '17:00'
    time_object_end_work = datetime.strptime(time_end_work_string, "%H:%M")  # date_object_dinner

    time_to_start_work_string = '8:00'
    time_object_to_start_work = datetime.strptime(time_to_start_work_string, "%H:%M")  # date_object_dinner

    delta_to_dinner = time_object_dinner - date_object_time_now               # time to dinner
    delta_to_end_work = time_object_end_work - date_object_time_now           # time to end work
    delta_to_start_work =  time_object_to_start_work - date_object_time_now   # time to start work

    engine.say('Сейчас {}'.format(time_now), 'report_current_time')

    if time_object_to_start_work > date_object_time_now:
       engine.say('До начала рабочего дня осталось {}. Натраиваемся на работу. '.format(delta_to_start_work), 'report_time_to_start_work')
    elif  time_object_dinner >= date_object_time_now >= time_object_to_start_work:
       engine.say('До обеда осталось {}'.format(delta_to_dinner), 'report_time_to_dinner')
       engine.say('До конца рабочего дня осталось {}'.format(delta_to_end_work), 'report_time_to_end_work')
    elif time_object_dinner <= date_object_time_now < time_object_dinner_end:
       engine.say('Обедаем и отдыхаем.'.format(delta_to_end_work), 'report_dinner')
       engine.say('До конца рабочего дня осталось {}'.format(delta_to_end_work), 'report_time_to_end_work')
    elif date_object_time_now <= time_object_end_work:
       engine.say('До конца рабочего дня осталось {}'.format(delta_to_end_work), 'report_time_to_end_work')
    elif date_object_time_now > time_object_end_work:
       engine.say('Пора домой, рабочий день окончен', 'report_finish')

    engine.runAndWait()

def f_create_db():
    global flag_base_factory, d_ostatky
    print("Function 'f_create_db()' is working..........")

    time_modification = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    start = monotonic()
    if flag_base_factory == False:  # 'flag_base_factory=False' for preventing rewriting 'base_factory.db' after WRONG connection to SQL-1C-BASE
        print("\033[0mConnection to SQL-1C-BASE was WRONG. 'base_factory.db' is OLD\033[0m ")
        return None
    # 'd_ostatky'
    #  00000013144 ['Полуфабрикаты', 'Пластина боковая ПлБ41Т.Д', '', -120, '', '', '', '']
    # d_ostatky {Code: [group, nomenklatura, quantity_mk, quantity_ceh, PRODUCT, OBOROTY='', ZAPUSK = '', SHTRIBS='', 'OSTATKY CEH'}

    # data_czeny = (datetime.datetime(4019, 11, 18, 0, 0), Decimal('243.12'), 'РозничнаяП2', '00000014861', 'Ячейка К2 (ККУ)')

    d_db_czeny = []

    # d_db_czeny = {Code: [group, nomenklatura, , quantity_mk, price, priceP, priceP1, priceP2, priceP3}
    # d_db_czeny = 00000013363 ['Стойка', 'Стойка центральная СЦ-20.50 Т', 51, '247.74', '205.62', '205.62', '210.54', 0]

    d_db_czeny = {i: [d_ostatky[i][0], d_ostatky[i][1],d_ostatky[i][2],  0, 0, 0, 0, 0] for i in d_ostatky}
    for i in data_czeny:
        for cz in d_db_czeny:
            if i[3] == cz:
                if i[2] == 'Розничные':
                    l_tmp = d_db_czeny[cz]
                    l_tmp[3] = '{:.2f}'.format(i[1])
                    d_db_czeny[cz] = l_tmp
                    break
                elif i[2] == 'РозничнаяП':
                    l_tmp = d_db_czeny[cz]
                    l_tmp[4] = '{:.2f}'.format(i[1])
                    d_db_czeny[cz] = l_tmp
                    break
                elif i[2] == 'РозничнаяП1':
                    l_tmp = d_db_czeny[cz]
                    l_tmp[5] = '{:.2f}'.format(i[1])
                    d_db_czeny[cz] = l_tmp
                    break
                elif i[2] == 'РозничнаяП2':
                    l_tmp = d_db_czeny[cz]
                    l_tmp[6] = '{:.2f}'.format(i[1])
                    d_db_czeny[cz] = l_tmp
                    break
                elif i[2] == 'РозничнаяП3':
                    l_tmp = d_db_czeny[cz]
                    l_tmp[7] = '{:.2f}'.format(i[1])
                    d_db_czeny[cz] = l_tmp
                else:
                    print('Unknown parameter = ', i[2])
    print("\033[0m'd_db_czeny'  is completed. \033[32m {:.5f} seconds \033[0m ".format(monotonic() - start))

    # creating "base_factory.db" if it does not exist
    if not os.path.exists('base_factory.db'):
        print("'base_factory.db' does not exist. We try to create it.")
        try:
            conn = sqlite3.connect('base_factory.db')
            c = conn.cursor()
            c.execute("CREATE TABLE prices_info (code INTEGER PRIMARY KEY NOT NULL, info VARCHAR(150))")
            conn.commit()

            c.execute("INSERT INTO prices_info VALUES (?, ?)", (1, 'first creation'))
            conn.commit()

            c.execute("INSERT INTO prices_info VALUES (?, ?)", (2, 'first creation'))
            conn.commit()

            # 'group' because of GROUP
            c.execute("CREATE TABLE prices   (code INTEGER PRIMARY KEY NOT NULL, \
               'group' VARCHAR(50) NOT NULL,\
               name VARCHAR(50) NOT NULL,\
               ostatok_mk INTEGER NOT NULL,\
               price NUMERIC(2) NOT NULL,\
               priceP NUMERIC(2) NOT NULL,\
               priceP1 NUMERIC(2) NOT NULL,\
               priceP2 NUMERIC(2) NOT NULL,\
               priceP3 NUMERIC(2) NOT NULL)")
            conn.commit()
        except Error as e:
            print('ERROR creation sql.db = {}'.format(e))
        finally:
            if conn:
                conn.close()

        print("'base_factory.db' was created")


    try:
        start = monotonic()
        conn = sqlite3.connect("base_factory.db")
        cursor = conn.cursor()

        # Deleting previous data
        cursor.execute("DELETE FROM prices")
        conn.commit()
        # cursor.execute("DELETE FROM prices_info")
        # conn.commit()

        # d_db_czeny = {Code: [group, nomenklatura, quantity_mk, price, priceP, priceP1, priceP2, priceP3}
        # d_db_czeny = 00000013363 ['Стойка', 'Стойка центральная СЦ-20.50 Т', 51, '247.74', '205.62', '205.62', '210.54', 0]
        # d_db_czeny is DICTIONARY. DATA from 'prices_info' wil be sorted 'ORDER BY' in APP (CLIENT_FACTORY)
        for i in d_db_czeny:
            group, name, ostatok_mk, price, priceP, priceP1, priceP2, priceP3 = d_db_czeny[i]
            if group != 'Полуфабрикаты':
                cursor.execute("INSERT INTO prices VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)",\
                               (i, group, name, ostatok_mk, price, priceP, priceP1, priceP2, priceP3))
        conn.commit()

        print("'base_factory.db' is modified at {}".format(time_modification))
        # cursor.execute("INSERT INTO prices_info VALUES(?, ?)", (1, time_modification))

        cursor.execute("UPDATE prices_info SET info='{}' WHERE code=1".format (time_modification))
        conn.commit()
        conn.close()
        print("\033[0m'base_factory' is completed. \033[32m {:.5f} seconds \033[0m ".format(monotonic() - start))
    except Exception as e:
        print("Exception: ", e)
        print("\033[32mCan not connect to intire DATABASE 'base_factory.db' \033[0m ")

    #loading 'base_factory.db' to FTP-server
    start = monotonic()
    print("start loading 'base_factory.db' on FTP-SERVER")
    try:
        ftps = FTP_TLS(helpic.place_ftp, helpic.user_ftp, helpic.password_ftp)
        print("Connection to FTP-SERVER is OK")
    except Exception as e:
        print(e)
        print('Connection', 'You don\'t have internet conection\n\
                                   or login/pasword were changed')
        return None  # for prevent doing next code

    ftps.cwd(helpic.directory_ftp)
    file_name = 'base_factory.db'
    ftps.storbinary('STOR ' + file_name, open('base_factory.db', '+rb'))  # загрузка файла НА сервер

    ftps.quit()
    # time_modification = 'base_factoty.db' last modification
    root.title("FACTORY (Release 1_0_March_2021)  'base_factory.db' was sent to FTP-Server at {}".format(time_modification))  # MAIN TITLE
    print("\033[0m 'base_factory.db' was uploaded on FTP-SERVER.  \033[32m {:.5f} seconds \033[0m ".format(monotonic() - start))

def f_open_rasschet():
    global tree_screen_rasschet , fra_rasschet, d_ostatky_copy_screen

    fra_rasschet = Toplevel()
    fra_rasschet.geometry('1800x900')
    fra_rasschet.title('RASSCHET')

    btn_shtribs_rasschet = MyButton(fra_rasschet, text='ШТРИБС')
    ent_length_shtribs = Entry(fra_rasschet,  width=5, font='arial 12', relief=SUNKEN, bd=2)
    ent_length_shtribs.focus()
    label_rasschet_product = MyLabel(fra_rasschet, text='ПРОДАЖИ')
    ent_count_rasschet = Entry(fra_rasschet, width=2, font='arial 12', relief=SUNKEN, bd=2)
    label_zapusk_length = MyLabel(fra_rasschet, text='ЗАПУСК ШТРИБС')
    ent_zapusk_shtribs = Entry(fra_rasschet, width=10, font='arial 12', relief=SUNKEN, bd=2)
    # btn_shtribs_save = MyButton(fra_rasschet, text='Сохранить', command=f_ostatky_shtribs)
    btn_shtribs_save = MyButton(fra_rasschet, text='Сохранить')

    ent_length_shtribs.focus()

    var_shtribs = IntVar()
    var_shtribs.set(0)
    rad_product = Radiobutton(fra_rasschet, text='Остатки Цех', variable=var, value=13)
    label_zapusk_need_length = MyLabel(fra_rasschet, text='Минимальный ЗАПУСК ')
    label_zapusk_1_length = MyLabel(fra_rasschet, text='ЗАПУСК ФАКТ ')

    fra_rasschet_main = MyLabelFrame(fra_rasschet, width=1200, height=700, text='RASSCHET', bg='bisque')

    btn_shtribs_rasschet.grid(row=1, column=1, padx=5, pady=5, sticky='w')
    ent_length_shtribs.grid(row=1, column=2, padx=5, pady=5, sticky='w')
    label_rasschet_product.grid(row=1, column=3, padx=5, pady=5, sticky='w')
    ent_count_rasschet.grid(row=1, column=4, padx=5, pady=5, sticky='w')
    label_zapusk_length.grid(row=1, column=5, padx=5, pady=5, sticky='w')
    ent_zapusk_shtribs.grid(row=1, column=6, padx=5, pady=5, sticky='w')
    btn_shtribs_save.grid(row=1, column=7, padx=5, pady=5, sticky='w')

    rad_product.grid(row=2, column=1, padx=5, pady=5, sticky='w')
    label_zapusk_need_length.grid(row=2, column=2, padx=5, pady=5, sticky='w')
    label_zapusk_1_length.grid(row=2, column=3, padx=5, pady=5, sticky='w')

    fra_rasschet_main.grid(row=3, column=0, columnspan=12)

    tree_screen_rasschet = ttk.Treeview(fra_rasschet_main)
    tree_screen_rasschet['height'] = 35
    tree_screen_rasschet['column'] = ('Штрибс', 'Длина', 'Ширина', "К-во 'X'", "К-во 'Y'",'М/к', 'Цех №2', 'Заявка',\
                                      'Остаток', 'Продажи' ,'K_до', 'ЗАПУСК', 'K_после')
    tree_screen_rasschet.column('#0', width=170, anchor='n')    # Is not WORKING !!!!!!!!
    tree_screen_rasschet.column('Штрибс', width=80, anchor='n')
    tree_screen_rasschet.column('Длина', width=80, anchor='n')
    tree_screen_rasschet.column('Ширина', width=80, anchor='n')
    tree_screen_rasschet.column("К-во 'X'", width=80, anchor='n')
    tree_screen_rasschet.column("К-во 'Y'", width=80, anchor='n')
    tree_screen_rasschet.column('М/к', width=80, anchor='n')
    tree_screen_rasschet.column('Цех №2', width=80, anchor='n')
    tree_screen_rasschet.column('Заявка', width=80, anchor='n')
    tree_screen_rasschet.column('Остаток', width=80, anchor='n')
    tree_screen_rasschet.column('Продажи', width=80, anchor='n')
    tree_screen_rasschet.column('K_до', width=80, anchor='n')
    tree_screen_rasschet.column('ЗАПУСК', width=80, anchor='n')
    tree_screen_rasschet.column('K_после', width=80, anchor='n')

    tree_screen_rasschet.heading('#0', text='Номенклатура')
    tree_screen_rasschet.heading('Штрибс', text='Штрибс')
    tree_screen_rasschet.heading('Длина', text='Длина')
    tree_screen_rasschet.heading('Ширина', text='Ширина')
    tree_screen_rasschet.heading("К-во 'X'", text="К-во 'X'")
    tree_screen_rasschet.heading("К-во 'Y'", text="К-во 'Y'")
    tree_screen_rasschet.heading('М/к', text='М/к')
    tree_screen_rasschet.heading('Цех №2', text='Цех №2')
    tree_screen_rasschet.heading('Заявка', text='Заявка')
    tree_screen_rasschet.heading('Остаток', text='Остаток')
    tree_screen_rasschet.heading('Продажи', text='Продажи')
    tree_screen_rasschet.heading('K_до', text='K_до')
    tree_screen_rasschet.heading('ЗАПУСК', text='ЗАПУСК')
    tree_screen_rasschet.heading('K_после', text='K_после')

    tree_screen_rasschet.grid(row=1, column=0)

    # for deliting row by DELETE. extra copy
    d_ostatky_copy_screen = copy.deepcopy(d_ostatky)

    btn_shtribs_save.bind('<ButtonRelease-1>', lambda event, f1=ent_length_shtribs, f2=ent_count_rasschet, f3=label_rasschet_product, \
                                     f4=ent_zapusk_shtribs, f5=label_zapusk_need_length, f6=label_zapusk_1_length: \
                                  f_ostatky_shtribs(event, f1.get(), f2.get(), f3, f4.get(), f5, f6))

    btn_shtribs_rasschet.bind('<Button-1>', lambda event, f1=ent_length_shtribs, f2=ent_count_rasschet, f3=label_rasschet_product,\
        f4=ent_zapusk_shtribs, f5=label_zapusk_need_length, f6=label_zapusk_1_length:\
        f_rasschet(event, f1.get(), f2.get(), f3, f4.get(), f5, f6))

    ent_length_shtribs.bind('<Return>',lambda event, f1=ent_length_shtribs, f2=ent_count_rasschet, f3=label_rasschet_product, \
                                     f4=ent_zapusk_shtribs, f5=label_zapusk_need_length, f6=label_zapusk_1_length: \
                                  f_rasschet(event, f1.get(), f2.get(), f3, f4.get(), f5, f6))

    ent_count_rasschet.bind('<Return>', lambda event, f1=ent_length_shtribs, f2=ent_count_rasschet, f3=label_rasschet_product, \
                                   f4=ent_zapusk_shtribs, f5=label_zapusk_need_length, f6=label_zapusk_1_length: \
                                f_rasschet(event, f1.get(), f2.get(), f3, f4.get(), f5, f6))

    ent_zapusk_shtribs.bind('<Return>',lambda event, f1=ent_length_shtribs, f2=ent_count_rasschet, f3=label_rasschet_product, \
                                   f4=ent_zapusk_shtribs, f5=label_zapusk_need_length, f6=label_zapusk_1_length: \
                                f_rasschet(event, f1.get(), f2.get(), f3, f4.get(), f5, f6))


    tree_screen_rasschet.bind('<Double-Button-1>',  lambda event, f1=ent_length_shtribs, f2=ent_count_rasschet, f3=label_rasschet_product,\
        f4=ent_zapusk_shtribs, f5=label_zapusk_need_length, f6=label_zapusk_1_length:\
        f_double_click_rasschet(event, f1.get(), f2.get(), f3, f4.get(), f5, f6))

    tree_screen_rasschet.bind('<Return>',lambda event, f1=ent_length_shtribs, f2=ent_count_rasschet, f3=label_rasschet_product, \
         f4=ent_zapusk_shtribs, f5=label_zapusk_need_length, f6=label_zapusk_1_length: \
         f_double_click_rasschet(event, f1.get(), f2.get(), f3, f4.get(), f5, f6))

    tree_screen_rasschet.bind('<Delete>', lambda event, f1=ent_length_shtribs, f2=ent_count_rasschet, f3=label_rasschet_product, \
                                    f4=ent_zapusk_shtribs, f5=label_zapusk_need_length, f6=label_zapusk_1_length: \
                                 f_delete_item_screen_rasschet(event, f1.get(), f2.get(), f3, f4.get(), f5, f6))

# creating table for RASSCHET
def f_rasschet(event, ent_length_shtribs, ent_count_rasschet, label_rasschet_product, ent_zapusk_shtribs, label_zapusk_need_length,\
               label_zapusk_1_length, focus_item=0):

    f_rasschet_func(ent_length_shtribs, ent_count_rasschet, label_rasschet_product, ent_zapusk_shtribs,
                    label_zapusk_need_length,label_zapusk_1_length, focus_item)


def f_rasschet_func(ent_length_shtribs, ent_count_rasschet, label_rasschet_product, ent_zapusk_shtribs, label_zapusk_need_length,\
               label_zapusk_1_length, focus_item):
    global d_ostatky, d_karta_detaley, d_zayavka, tree_screen_rasschet, l_for_save, d_ostatky_copy_screen, l_position

    ent_zapusk_shtribs = 0 if not ent_zapusk_shtribs else int(ent_zapusk_shtribs)
    # 'd_ostatky'
    #  00000013144 ['Полуфабрикаты', 'Пластина боковая ПлБ41Т.Д', '', -120, '', '', '', '']
    # d_ostatky {Code: [group, nomenklatura, quantity_mk, quantity_ceh, PRODUCT, OBOROTY='', ZAPUSK = '', SHTRIBS='', 'OSTATKY CEH'}
    #                                             width length   shtribs   x     y

    # d_karta_detaley  {'Б 75 ЕР.Д': ['kharkov', '159', '67',    '159',   '14', '1']}

    # deleting rows before laoding data
    rows = tree_screen_rasschet.get_children()
    for item in rows:
        tree_screen_rasschet.delete(item)

    # d_ostatky_div_shtribs = {'00000011725': ['Полуфабрикаты', 'ПБС 600 ЕР.Поперечина боковой стойки.Д', '', 223, '', 1636, '', '138', 0],
    d_ostatky_div_shtribs = copy.deepcopy(d_ostatky_copy_screen)  # DEEP COPY !!!!!!!

    for i in d_ostatky_div_shtribs:
        d_ostatky_div_shtribs [i] = d_ostatky_div_shtribs [i] + [i]   # for ZAYAVKA
    # 00000012575 ['МКС ', 'МКС 1063 IP54 (без МП) ', '', '', '', 6, '', '', '', '00000012575']

    try:
        divisior = int(ent_count_rasschet) if int(ent_count_rasschet) > 0 else 1
    except ValueError:
        divisior = 1

    flag = True if (ent_length_shtribs and ent_length_shtribs != '0')  else False
    counter = 1
    l_for_save = []
    full_length_zapusk = 0
    zapusk_1_length = 0

    if flag:   #shtribs in window
        d_zapusk_shtribs = {}
        for i in d_ostatky_div_shtribs.values():
            if i[1] in d_karta_detaley and d_karta_detaley.get(i[1], None)[3] == ent_length_shtribs:
                # tree_screen_rasschet['column'] = ('Длина', "К-во 'X'", "К-во 'Y'", 'М/к', 'Цех №2', 'Заявка', 'Остаток', 'Продажи', 'Запуск')  !!! to change in the end
                l = d_karta_detaley.get(i[1], None)
                width, length, shtribs, x, y = l[1], l[2], l[3], l[4], l[5]
                length = ''.join(length.split())                 # 1 098  !!!!!
                ostatok_mk = 0 if i[2] == '' else int(i[2])
                ostatok_ceh = 0 if i[3] == '' else int(i[3])
                zayavka = sum([int(iii[2]) for iii in d_zayavka.get(i[9], None)]) if d_zayavka.get(i[9], None) != None else 0
                ostatok = ostatok_mk + ostatok_ceh - zayavka

                zayavka = zayavka if zayavka else ''
                product = round(int(i[5] if i[5] else 0)/divisior)

                zapusk_min = product - ostatok

                full_length_zapusk += (zapusk_min if zapusk_min > 0 else 0)*int(length)/int(y)

                if product != 0:
                    percent_before = round(ostatok / product, 2) if product else 0  # for situation where product = 0   'ZeroDivision'
                else:
                    percent_before = 9999

                percent_tmp = percent_before
                zapusk = 0
                d_zapusk_shtribs[i[1]] = [shtribs, length, width, x, y, i[2], i[3], zayavka, ostatok, product, percent_before, percent_tmp, zapusk]

        while zapusk_1_length < int(ent_zapusk_shtribs) * 1000:
            if len(d_zapusk_shtribs) == 0:   # preventing doing next code if table is EMPTY
                print("No items for calculation......")
                return 1
            item_min = min(d_zapusk_shtribs, key=lambda x: d_zapusk_shtribs[x][11])    # key=lambda x    =  percent_tmp
            l_tmp = d_zapusk_shtribs[item_min]
            l_tmp[12] += int(l_tmp[4])                                      # zapusk # l_tmp[4] = y

            l_tmp[11] = round((l_tmp[8]+ l_tmp[12])  / l_tmp[9], 2)         # percent_tmp # zapusk = l_tmp[11],  ostatok = l_tmp[7], product = l_tmp[8]

            d_zapusk_shtribs[item_min] = l_tmp
            zapusk_1_length += int(l_tmp[1])                                # length = l_tmp[1]

        l_position = [(num, i) for num, i in enumerate(sorted(d_zapusk_shtribs))]  #for getting position DELETE
        # d_zapusk_shtribs = {'Б 75 ЕР.Д': ['159', '67', '159', '14', '1', '', 1094, '',  1094, 11440, 0.1, 0.1, 0]
        # for items in sorted(d_zapusk_shtribs, key=lambda x: d_zapusk_shtribs[x][10]):          # sort by 'k_do'
        for num, items in l_position:                                                   # sort by 'nomenklatura'
            shtribs, length, width, x, y, ost_mk, ost_ceh, zayavka, ostatok, product, percent_before, percent_tmp, zapusk = d_zapusk_shtribs[items]
            if product != 0:
                percent_after =  round((zapusk + ostatok) / product, 2) if  ent_zapusk_shtribs else ''
            else:
                percent_after = ''
            zapusk = zapusk if zapusk else ''

            tree_screen_rasschet.insert('', counter, text=items, values=(shtribs, length, width, x, y, ost_mk, ost_ceh, zayavka, ostatok, product, percent_before, zapusk, percent_after))
            counter += 1
            # for preventing saving items without  'zapusk'
            if zapusk:
                l_for_save.append([items, shtribs, zapusk])

    else:
        l_position = []
        # d_ostatky_div_shtribs = {'00000009751': ['Полуфабрикаты', 'Стойка вертикальная СВ ТС 1800', '', 69, '', 208, '', '146', '', '00000009751'],
        for i in sorted(d_ostatky_div_shtribs.values(), key=lambda x: x[1]):       # sorted by name nomenklatura
            # i[7] != '0'     shtribs
            if i[1] in d_karta_detaley and d_karta_detaley.get(i[1], None)[3] != '' and i[7] != '0':
                l_position.append(i[1])
                # tree_screen_rasschet['column'] = ('Длина', "К-во 'X'", "К-во 'Y'", 'М/к', 'Цех №2', 'Заявка', 'Остаток', 'Продажи', 'Запуск')
                l = d_karta_detaley.get(i[1], None)
                width, length, shtribs, x, y = l[1], l[2], l[3], l[4], l[5]
                length = ''.join(length.split())                   # 1 098  !!!!!
                ostatok_mk  = 0 if i[2] == '' else int(i[2])
                ostatok_ceh = 0 if i[3] == '' else int(i[3])
                zayavka = sum([int(iii[2]) for iii in d_zayavka.get(i[9], None)]) if d_zayavka.get(i[9], None) != None else 0
                ostatok = ostatok_mk + ostatok_ceh - zayavka
                zayavka = zayavka if zayavka else ''
                product = round(int(i[5] if i[5] else 0)/divisior)

                percent_before = round(ostatok / product, 2) if product else 0  # for situation where product = 0

                tree_screen_rasschet.insert('', counter, text=i[1], values=(shtribs, length, width,  x, y, i[2],\
                                                                            i[3], zayavka, ostatok, product, percent_before))
                counter += 1
        l_position = list(enumerate(l_position))    #rewriting 'l_position' like for situation with flag
    try:
        item_selected = tree_screen_rasschet.get_children()[focus_item]
        tree_screen_rasschet.selection_set(item_selected)  # for lighting  position in the row
        tree_screen_rasschet.focus_set()                   # for keys UP and DOWN
        tree_screen_rasschet.focus(item_selected)          # for keys UP and DOWN
    except IndexError:
        print('You delete ALL rows')

    label_rasschet_product.configure(text='ПРОДАЖИ /{}'.format(divisior))
    label_zapusk_need_length.configure(text='Минимальный ЗАПУСК {}'.format(round(full_length_zapusk)))
    label_zapusk_1_length.configure(text="ЗАПУСК ФАКТ {}".format(zapusk_1_length if zapusk_1_length != 0 else ''))


# writing info to 'ostatky_shtribs.txt' from 'screen_rasschet'
def f_ostatky_shtribs(event, ent_length_shtribs, ent_count_rasschet, label_rasschet_product, ent_zapusk_shtribs, label_zapusk_need_length,\
               label_zapusk_1_length, focus_item=0):
    global l_for_save

    if not l_for_save:
        showinfo('SaveZapusk', 'Вы забыли создать ЗАПУСК')
        return 1
    # l_for_save = [['Вставка Опоры ВО 100.Д', '159', 798], ['Основание Опоры.ОО100.Д', '159', 336], ['Б 75 ЕР.Д', '159', ''],
    shtribs = l_for_save[0][1]

    #creating 'ostatky_shtribs.txt' if it does not exist
    if not os.path.exists('ostatky_shtribs.txt'):
        with open('ostatky_shtribs.txt', 'w', encoding='utf-8') as f:
            f.write('')
            print("'ostatky_shtribs.txt' was created")

    fra_rasschet.grab_set()

    if askyesno('Save zapusk shtribs','{:^25}\n{:^25}\n{:^25}'.format('Вы хотите сохранить', 'ЗАПУСК ШТРИБС', 'в файл?')):
        with open('ostatky_shtribs.txt', 'r', encoding='utf-8') as f:
             data = f.readlines()

        s_previous = ''.join(data)

        s = ''
        s += 'ЗАПУСК на дату: {} ШТРИБС: {}\n'.format(f_time_now(), shtribs)
        counter = 1
        for i in l_for_save:
            if i[2]:  # only items with quantity
                s += '{:>2} {:<40.40} {:>8} шт.\n'. format(counter, i[0], i[2])
            counter += 1
        s += '* ' * 30
        s += '\n{}'.format(s_previous)

        with open('ostatky_shtribs.txt', 'w', encoding='utf-8') as f:
             f.write(s)

        showinfo('ЗАПУСК ШТРИБС', '{:^25}\n{:^25}'.format('Информация сохранена в' , 'ostatky_shtribs.txt'))

        # fra_rasschet.destroy()

        # Renew FRAME RASSCHET

        f_rasschet_func(ent_length_shtribs, ent_count_rasschet, label_rasschet_product, ent_zapusk_shtribs,
                    label_zapusk_need_length,label_zapusk_1_length, focus_item=0)


def f_karta_detaley():
    global d_karta_detaley, data_karta_detaley, d_original, d_original
    # d_original = {'00000013791': ['ЕР 16104/2В', 'ЕР '],
    #                            width             length          shtribs             X             Y           place             code              nomenklatura
    # data_karta_detaley = (Decimal('329.00'), Decimal('1113.00'), Decimal('329'), Decimal('1'), Decimal('1'), 'хр          ', '00000011317', 'ОВ653М.Обечайка верхняя.Д')
    #                   {nomenklatura: [place,     width, length,   shtribs,  x,    y]}
    # d_karta_detaley = {'Б 75 ЕР.Д':  ['kharkov', '159', '67',  '159',   '14', '1'],
    d_karta_detaley = {}
    for i in data_karta_detaley:
            if i[6] in d_original:   # code
                nomenklatura = i[7].strip()
                try:                      # because of empty lines
                    place        = 'kharkov' if i[5].startswith('хр') else 'work'
                except AttributeError:
                    print (i, i[5] )
                    break

                length       = str(int(i[1]))
                width        = str(int(i[0]))
                shtribs      = str(i[2])
                x            = str(i[3]) if i[3] else ''
                y            = str(i[4]) if i[4] else ''
                d_karta_detaley[nomenklatura] = [place,  width, length,  shtribs, x, y]

# statistik for eleton.herokuapp.com
def f_statistica():
    start = monotonic()
    # connection to FTP-Server
    try:
        ftps = FTP_TLS(helpic.place_ftp, helpic.user_ftp, helpic.password_ftp)
        print("Connection to FTP-SERVER is OK")
    except Exception as e:
        print(e)
        print('Connection', 'You don\'t have internet conection or login/pasword were changed')
    # checking FOLDER at FTP-Server
    try:
        ftps.cwd(helpic.directory_ftp)
    except Exception as e:
        print(e)
        print("\033[31mCan not find directory '{}' at FTP-Server\033[0m".format(helpic.directory_ftp))
        return None  # for preventing doing next code

    # 'info_connection.db' if it does not exist on LOCAL folder
    if not os.path.exists('info_connection.db'):
       print("file 'info_connection.db' will be created")

    # checking 'info_connection.db' at FTP-Server
    try:
        file_name = 'info_connection.db'
        # 1. Copy 'info_connection.db' from FTP-Server to MAIN APP
        with open(file_name, 'wb') as f:
            ftps.retrbinary('RETR ' + file_name, f.write)     # rewriting 'info_connection.db' in LOCAL folder from FTP

    except Exception as e:
        print(e)
        print("\033[31mCan not find 'info_connection.db' in directory '{}' at FTP-Server\033[0m".format(helpic.directory_ftp))
        return None    # for preventing doing next code

    #copy files from FTP to LOCAL folder and removing them from FTP
    l_files_ftp = ftps.nlst()
    # We are looking for files like 'info_connection_heroku_20210223141959182592.db'

    for file_name_ftp in l_files_ftp:
        if file_name_ftp.startswith('info_connection_heroku'):
            with open(file_name_ftp, 'wb') as f:
                ftps.retrbinary('RETR ' + file_name_ftp, f.write)  # rewriting 'info_connection.db' in LOCAL folder from FTP
            ftps.delete(file_name_ftp)

    # getting info from 'info_connection.db' at LOCAL
    conn = sqlite3.connect('info_connection.db')
    cursor = conn.cursor()
    # 2021-02-12 18:05:59:011026 == [(95, '20210205113224719863', '2021-02-12 18:05:59:011026', '127.0.0.1', '49545', 'index_page')]
    data_new_all = {i[2] : [i] for i in cursor.execute("SELECT * FROM heroku_actions").fetchall()}
    conn.close()

    # getting info from files like 'info_connection_heroku_20210223141959182592.db'
    l_files_local = [i for i in os.listdir() if i.startswith('info_connection_heroku')]
    for file_name_local in l_files_local:
        conn_1 = sqlite3.connect(file_name_local)
        cursor_1 = conn_1.cursor()
        try:
            data_new = cursor_1.execute("SELECT * FROM heroku_actions")
            data_new = data_new.fetchall()
            conn_1.close()
            for i in data_new:
                if i[2] not in data_new_all and  i[3]!='127.0.0.1':  # '127.0.0.1' - localhost for testing
                    # New record 2021 - 05 - 21 11: 06:32: 540600 === (2644, '20210217103455624309', '2021-05-21 11:06:32:540600', '127.0.0.1', 'index_page')
                    print("New record ", i[2], '===', i)
                    data_new_all [i[2]] = [i]
        except sqlite3.DatabaseError:
            print('This db is broken: {}'.format(file_name_local))
            continue

    # add NEW info
    print("Rewriting new info in 'info_connection.db'")
    #[(2644, '20210217103455624309', '2021-05-21 11:06:32:540600', '127.0.0.1', 'index_page')]
    data_for_db = [i for i in data_new_all.values() if i[0][1]]       # preventing None === (0, None, None, None, None)
    data_for_db.sort(key=lambda x: x[0][2])
    conn = sqlite3.connect('info_connection.db')
    cursor = conn.cursor()
    cursor.execute("DELETE FROM heroku_actions")
    conn.commit()

    for i in data_for_db:
        cursor.execute("INSERT INTO heroku_actions VALUES (?, ?, ?, ?, ?)", (None, i[0][1], i[0][2], i[0][3], i[0][4]))
    conn.commit()
    conn.close()

    # Rewriting file 'info_connection.db' on FTP server
    print("'info_connection.db': ", 'info_connection.db')
    ftps.storbinary('STOR ' + file_name, open(file_name, '+rb'))  # загрузка файла НА FTP-Server

    ftps.quit()
    print("Information in 'info_connection.db' WAS UPDATED ON FTP-Server")

    # creating EXcell file
    conn = sqlite3.connect('info_connection.db')
    cursor = conn.cursor()
    # for situation, when something wrong with 'info_connection.db'
    try:
        data = cursor.execute("SELECT * FROM heroku_actions")
    except (sqlite3.OperationalError, sqlite3.DatabaseError) as er:
        print("\033[31m We have some problem with database 'info_connection.db'.....\033[0m", er)
        return None
    data = data.fetchall()

    # cursor.description - list with column's names
    columns = [desc[0] for desc in cursor.description]     # for Excell-file
    conn.close()

    wb = Workbook()
    file_xlsx_path = 'statistic.xlsx'

    try:
        wb.save(file_xlsx_path)

        for file_name_local in l_files_local:  # deleating files only when recording to 'statistic.xlsx' is possible
            os.remove(file_name_local)

    except PermissionError:
        print("Can not write 'statistic.xlsx'. File is opened now. First saving.")
        return 1

    sheet = wb.active

    counter_col = 1
    counter_row = 2
    for i in columns:
        sheet.cell(row=1, column=counter_col).value = i
        counter_col += 1

    # [(95, '20210205113224719863', '2021-02-12 18:05:59:011026', '127.0.0.1', 'index_page')]
    for i in data_for_db:
        code, identity, date_time, ip_address, action = i[0]

        sheet.cell(row=counter_row, column=1).value = code
        sheet.cell(row=counter_row, column=2).value = identity
        sheet.cell(row=counter_row, column=3).value = date_time
        sheet.cell(row=counter_row, column=4).value = ip_address
        sheet.cell(row=counter_row, column=5).value = action

        counter_row += 1

    try:
        wb.save(file_xlsx_path)

    except PermissionError:
        print("Can not write 'statistic.xlsx'. File is opened now. Second saving")
        return 2

    print("\033[0m 'statistic.xlsx' was updated from FTP-SERVER.  \033[32m {:.5f} seconds \033[0m ".format(monotonic() - start))

# BUTTON 'СTАТИСТИКА'
def f_heroku():
    f_statistica()
    showinfo('HEROKU', "'statistic.xls' WAS UPDATED")
    return 0

# Field MESSAGE
def f_message_check():
    if os.path.exists("base_factory.db"):
        try:
            conn = sqlite3.connect("base_factory.db")
            cursor = conn.cursor()

            data = cursor.execute("SELECT info FROM prices_info WHERE code=2")
            data = data.fetchone()
            # data =  ('qqqq\n',)
            data = data[0].strip()
            conn.close()
            text_message.delete(1.0, END)
            text_message.insert(INSERT, data)
        except Exception as e:
            print("f_message_check(). Can not GET check new MESSAGE from 'base_factory.db'")
            print(e)
        return 0
    print("f_message_check(). Can not check new MESSAGE from 'base_factory.db'")
    return 1

# button MESSAGE
def f_message_put():
    if os.path.exists("base_factory.db"):
        res = text_message.get(1.0, END)
        conn = sqlite3.connect("base_factory.db")
        cursor = conn.cursor()
        cursor.execute("UPDATE prices_info SET info='{}' WHERE code=2".format(res))
        conn.commit()
        conn.close()
    print("f_message_put(). Can not put new MESSAGE to 'base_factory.db'")

# for WORK and DEMO versions
def f_main():
    global d_ostatky, d_zayavka, d_rezerv, d_karta_detaley, d_union_sel, data_shtribs
    if flag_base_factory:              # if connection to 1C SERVER is OK
        f_karta_detaley()              # creating d_karta_detaley
        f_ostatky_ceh_mk()             # 'd_ostatky' with 'ostatki mk', 'ostatki ceh'
        f_oboroty()                    # adding 'info oboroti' to 'd_ostatky'
        f_nezavershonnie_naryadi()     # creating 'd_union_sel' = nezavershonnie_naryadi
        f_zayavka()                    # creating 'd_zayavka' = zayavki pokupateley
        f_rezerv()                     # creating 'd_rezerv' = rezetvi pokupateley
        f_add_pr_st_zapusk()           # adding info from 'd_union_sel' to 'd_ostatky'

    else:                              # for DEMO version ONLY
        print('Trying to represent DEMO-version')
        d_ostatky       = helpic.d_ostatky
        d_zayavka       = helpic.d_zayavka
        d_rezerv        = helpic.d_rezerv
        d_karta_detaley = helpic.d_karta_detaley
        d_union_sel     = helpic.d_union_sel
        data_shtribs    = helpic.data_shtribs

        root.title('!!! DEMO. It is only DEMO version !!!')

        showinfo('DEMO', 'It is only DEMO version')
    return 2

# common traces for BUTTONs
class MyButton(Button):
    def __init__(self, *args, **kwargs):
        Button.__init__(self, *args, **kwargs)
        self['bg']               = 'lightgreen'
        self['activebackground'] = 'lightblue'
        self['fg']               = 'black'
        self['font']             = 'arial 12'
        self['relief']           = RAISED
        self['padx']             = 5
        self['pady']             = 5
        self['bd']               = 2

# common traces for LABELs
class MyLabel(Label):
    def __init__(self, *args, **kwargs):
        Label.__init__(self, *args, **kwargs)
        self['bg']               = 'lightblue'
        self['fg']               = 'black'
        self['font']             = 'arial 12'
        self['relief']           = FLAT
        self['padx']             = 5
        self['pady']             = 5
        self['bd']               = 2

# common traces for LabelFrame
class MyLabelFrame(LabelFrame):
    def __init__(self, *args, **kwargs):
        LabelFrame.__init__(self, *args, **kwargs)
        self['relief'] = RAISED
        self['padx'] = 5
        self['pady'] = 5
        self['bd'] = 5

root = Tk()
# root.title("FACTORY")
root.geometry('1950x900')

m = Menu()
root.config(menu=m)
fm = Menu(m, fg='green', font='arial 10')

m.add_cascade(label='File', menu=fm)
fm.add_command(label='About', command=f_about)
fm.add_command(label='Exit', command=f_exit)

btn_count = MyButton(root, text='Рассчет', command=f_store)
label_period_product = MyLabel(root, text='ПРОДАЖИ:{} - {}')

var = IntVar()
var.set(3)
rad_zayavka        = Radiobutton(root, text='Заявка', variable=var, value=9)
rad_rezerv         = Radiobutton(root, text='Резерв', variable=var, value=10)
rad_ostatok_ceh    = Radiobutton(root, text='Остаток Цех', variable=var, value=11)
rad_product        = Radiobutton(root, text='Производство', variable=var, value=1)
rad_specif_inf     = Radiobutton(root, text='Спецификация', variable=var, value=2)
rad_zapusk_inf     = Radiobutton(root, text='Запуск', variable=var, value=3)
rad_quantity       = Radiobutton(root, text='Колличество', variable=var, value=4)
rad_specif_middle  = Radiobutton(root, text='Спецификация', variable=var, value=5)
rad_ostatok_bottom = Radiobutton(root, text='Остаток Цех', variable=var, value=6)
rad_zapusk_bottom  = Radiobutton(root, text='Запуск', variable=var, value=7)
rad_product_bottom = Radiobutton(root, text='Производство', variable=var, value=8)


btn_zagotovka = MyButton(root, text='Заготовка', command=f_zagotovka_zapusk)

ent_count = Entry(root, width=2, font='arial 12', relief=SUNKEN, bd=2)
ent_count.focus()                    # focus for first view
ent_count.focus_set()                # focus for typing
ent_count.bind('<Return>', f_store_ent)

fra_top   = MyLabelFrame(root, width= 900, height=300, text='Information for MASTER', bg='lightgreen', padx=5, pady=5, relief=RAISED, bd=5)

fra_middle     = MyLabelFrame(root, width= 900, height=300, text='Selected items', bg='lightblue')
fra_store      = MyLabelFrame(root, width= 300, height=300, text='Previous zapusk', bg='lightblue')
fra_bottom     = MyLabelFrame(root, width= 1700, height=300, text='Production', bg='bisque')
btn_file       = MyButton(root, text='Сохранить в файл', command=f_save_file)
btn_delta      = MyButton(root, text='Сохранить DELTA', command=f_save_delta)
btn_screen     = MyButton(root, text='SCREEN', command=f_open_screen)
btn_save_store = MyButton(root, text='Сохранить', command=f_save_store)
btn_kharkov    = MyButton(root, text='Остатки Харьков', command=f_ostatky_kharkov)

fra_shtribs  = MyLabelFrame(root, width= 300, height=300, text='Shtribs', bg='bisque')
btn_product  = MyButton(root, text='ПРОИЗВОДСТВО', command=f_ostatky_product)
btn_refresh  = MyButton(root, text=text_refresh, command=f_refresh)
btn_shtribs  = MyButton(root, text='ШТРИБСЫ', command=f_shtribs)
btn_rasschet = MyButton(root, text='РАССЧЕТ', command=f_open_rasschet)
btn_report   = MyButton(root, text='ДОКЛАД', command=f_report)
btn_heroku   = MyButton(root, text='СTАТИСТИКА', command=f_heroku)
btn_message_put  = MyButton(root, text='M PUT', command=f_message_put)

text_message = Text(root, width=70, height=2, font='arial 12', relief=SUNKEN, bd=2)
try:
    f_message_check()
except sqlite3.OperationalError:
    print ("Can not find table 'prices_info'. Can not put TEXT in field ")

btn_message_check  = MyButton(root, text='M CHECK', command=f_message_check)

#Fonts for tree_inf
style = ttk.Style()
style.configure('Treeview', font='arial 12')
style.configure('Treeview.Heading', font='arial 12')

#Tree_TOP
tree_inf = ttk.Treeview(fra_top)
tree_inf['column'] = ('склад М/к', 'склад Цех', 'Остаток Цех', 'Производство', 'Остаток', 'ПРОДАЖИ', 'Заявка', 'Резерв', 'ЗАПУСК')
tree_inf.column('склад М/к', width=100,anchor='n')
tree_inf.column('склад Цех', width=100,anchor='n')
tree_inf.column('Остаток Цех', width=100,anchor='n')
tree_inf.column('Производство', width=100,anchor='n')
tree_inf.column('Остаток', width=100,anchor='n')
tree_inf.column('ПРОДАЖИ', width=100,anchor='n')
tree_inf.column('Заявка', width=100,anchor='n')
tree_inf.column('Резерв', width=100,anchor='n')
tree_inf.column('ЗАПУСК', width=100,anchor='n')
tree_inf.heading('склад М/к',text='склад М/к')
tree_inf.heading('склад Цех', text='склад Цех')
tree_inf.heading('Остаток Цех', text='Остаток Цех')
tree_inf.heading('Производство', text='Производство')
tree_inf.heading('Остаток', text='Остаток')
tree_inf.heading('ПРОДАЖИ', text='ПРОДАЖИ')
tree_inf.heading('Заявка', text='Заявка')
tree_inf.heading('Резерв', text='Резерв')
tree_inf.heading('ЗАПУСК', text='ЗАПУСК')
tree_inf.grid(row=1, column=0)

# Tree MIDDLE
tree_middle = ttk.Treeview(fra_middle)
tree_middle['column'] = ('К-во')
tree_middle.column('К-во', width=100,anchor='n')
tree_middle.heading('К-во', text='К-во')
tree_middle.grid(row=1, column=0)

# Tree STORE
tree_store = ttk.Treeview(fra_store)
tree_store['column'] = ('Дата')
tree_store.column('Дата', width=100,anchor='n')
tree_store.heading('Дата', text='Дата')
tree_store.grid(row=1, column=0)

# Tree BOTTOM
tree_bottom = ttk.Treeview(fra_bottom)
tree_bottom['height'] = 20
tree_bottom['column'] = ('ostatok_ceh', 'ostatok', 'Производство', 'need', 'delta', 'zapusk', 'width', 'long', 'shtribs', 'x', 'y', 'LONG','DELTA', 'zapusk_karta')
tree_bottom.column('ostatok_ceh', width=120, anchor='n')
tree_bottom.column('ostatok', width=80, anchor='n')
tree_bottom.column('Производство', width=150, anchor='n')
tree_bottom.column('need', width=80, anchor='n')
tree_bottom.column('delta', width=80, anchor='n')
tree_bottom.column('zapusk', width=80, anchor='n')
tree_bottom.column('width', width=80, anchor='n')
tree_bottom.column('long', width=80, anchor='n')
tree_bottom.column('shtribs', width=80, anchor='n')
tree_bottom.column('x', width=80, anchor='n')
tree_bottom.column('y', width=80, anchor='n')
tree_bottom.column('LONG', width=100, anchor='n')
tree_bottom.column('DELTA', width=100, anchor='n')
tree_bottom.column('zapusk_karta', width=100, anchor='n')
tree_bottom.heading('ostatok_ceh', text='Цех')
tree_bottom.heading('ostatok', text='Остаток Цех')
tree_bottom.heading('Производство', text='Производство')
tree_bottom.heading('need', text='НУЖНО')
tree_bottom.heading('delta', text='ДЕЛЬТА')
tree_bottom.heading('zapusk', text='ЗАПУСК')
tree_bottom.heading('width', text='ширина')
tree_bottom.heading('long', text='длина')
tree_bottom.heading('shtribs', text='штрибс')
tree_bottom.heading('x', text='x')
tree_bottom.heading('y', text='y')
tree_bottom.heading('LONG', text='ДЛИНА')
tree_bottom.heading('DELTA', text='DELTA')
tree_bottom.heading('zapusk_karta', text='Запуск_карта')
tree_bottom.grid(row=1, column=0)

tree_shtribs = ttk.Treeview(fra_shtribs)
tree_shtribs['column'] = ('quantity')
tree_shtribs.column('quantity', width=100, anchor='n')
tree_shtribs.heading('quantity', text='Длина, м')
tree_shtribs.grid(row=1, column=0)

btn_refresh.grid(row=1, column=0, padx=5, pady=5, sticky='w')
label_period_product.grid(row=1, column=1, padx=5, pady=5, sticky='w')
btn_count.grid(row=1, column=2, padx=5, pady=5, sticky='e')
ent_count.grid(row=1, column=3, padx=5, pady=5, sticky='w')
btn_product.grid(row=1, column=4, padx=5, pady=5, sticky='w')
btn_shtribs.grid(row=1, column=5, padx=5, pady=5, sticky='w')
btn_rasschet.grid(row=1, column=6, padx=5, pady=5, sticky='w')
# rad_ostatok_ceh.grid(row=2, column=0, padx=5, pady=5, sticky='w')
rad_zayavka.grid(row=2, column=1, padx=5, pady=5, sticky='w')
rad_rezerv.grid(row=2, column=2, padx=5, pady=5, sticky='w')
rad_product.grid(row=2, column=3, padx=5, pady=5, sticky='w')
rad_specif_inf.grid(row=2, column=4, padx=5, pady=5, sticky='w')
rad_zapusk_inf.grid(row=2, column=5, padx=5, pady=5, sticky='w')
rad_quantity.grid(row=2, column=6, padx=5, pady=5, sticky='e')
rad_specif_middle.grid(row=2, column=7, padx=5, pady=5, sticky='e')
fra_top.grid(row=3, column=0, columnspan=6, padx=5, pady=5)
fra_middle.grid(row=3, column=6, columnspan=8, padx=5, pady=5, sticky='w')
fra_store.grid(row=3, column=8, padx=5, pady=5, sticky='w')
rad_ostatok_bottom.grid(row=4, column=0, padx=5, pady=5, sticky='w')
rad_zapusk_bottom.grid(row=4, column=1, padx=5, pady=5, sticky='w')
rad_product_bottom.grid(row=4, column=2, padx=5, pady=5, sticky='w')
btn_file.grid(row=4, column=4, padx=5, pady=5, sticky='w')
btn_delta.grid(row=4, column=4, padx=5, pady=5, sticky='w')
btn_screen.grid(row=4, column=5, padx=5, pady=5, sticky='w')
btn_zagotovka.grid(row=4, column=6, padx=5, pady=5, sticky='w')
btn_save_store.grid(row=4, column=7, padx=5, pady=5, sticky='w')
btn_kharkov.grid(row=4, column=8, padx=5, pady=5, sticky='w')
fra_bottom.grid(row=5, column=0, columnspan=8, padx=5, pady=5, sticky='w')
fra_shtribs.grid(row=5, column=8, padx=5, pady=5, sticky='w')
btn_report.grid(row=6, column=0, padx=5, pady=5, sticky='w')
btn_heroku.grid(row=6, column=1, padx=5, pady=5, sticky='w')
btn_message_put.grid(row=6, column=2, padx=5, pady=5, sticky='w')
text_message.grid(row=6, column=3, columnspan=7, padx=5, pady=5, sticky='w')
btn_message_check.grid(row=6, column=7, padx=5, pady=5, sticky='w')

# functions for start
f_connection()              # SQL-quary to 1C
f_main()                    # for 'WORK' and 'DEMO' versions

f_load_delta()              # 'ostatky_kharkov.pkl' - load data
f_table_tree_inf()          # creating MAIN TABLE // 'table_tree_inf'
previous_zapusk()           # for loading data in table 'previous zapusk'
f_create_db()               # rewriting 'base_factory.db' and sending to FTP-SERVER
f_statistica()              # rewriting 'statistic.xls' from FTP-SERVER


print('finish ........................')

root.bind('<Control-z>', f_close)
tree_inf.bind('<Double-Button-1>', f_double_click_inf)
tree_inf.bind('<Return>', f_double_click_inf)
tree_middle.bind('<Double-Button-1>', f_double_click_middle)
tree_middle.bind('<Delete>', f_del_middle)
tree_bottom.bind('<Double-Button-1>', f_double_click_bottom)
tree_bottom.bind('<Delete>', f_delete_item)
tree_store.bind('<Double-Button-1>', f_input)
tree_store.bind('<Delete>', f_delete_item_store)

# for table SCREEN
var_screen = IntVar()
var_screen.set(13)              # creating SCREEN table by default = ALL PRODUCTION
var_screen_z = IntVar()
var_screen_z.set(11)

root.mainloop()

#hidden import !!!!!!
# pyinstaller --hidden-import=pyttsx3.drivers --hidden-import=pyttsx3.drivers.dummy --hidden-import=pyttsx3.drivers.espeak --hidden-import=pyttsx3.drivers.nsss --hidden-import=pyttsx3.drivers.sapi5 -F factory.py
